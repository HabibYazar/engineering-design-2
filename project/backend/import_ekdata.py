"""`data/ekdata` altındaki EK GERÇEK VERİ kümelerini üretim şemasına aktarır.

===========================================================================
KAPSAM
------
Klasör ÖZYİNELEMELİ taranır; dosya türü içeriğinden tanınır. Şu an tanınan
kümeler:

    ankara_bilim_yks_4year.csv        ÖSYM/YKS yerleştirme, 2022-2025
    ankara_bilim_yks_4year.json       (aynı verinin JSON kopyası)
    ankara_bilim_yks_program_summary.csv   türetilmiş özet
    ankara_bilim_mufredat_data1.xlsx  müfredat / ders kataloğu

Tanınmayan dosya SESSİZCE ATLANMAZ; raporun sonunda listelenir.

YÖK VERİSİYLE BİRLİKTE ÇALIŞIR
------------------------------
Bu betik YÖK Akademik verisini SİLMEZ ve üzerine yazmaz. Aynı üniversiteyi
tamamlar: YÖK akademik kadroyu, ekdata öğrenci yerleştirmesini ve müfredatı
getirir. İkisi aynı fakülte/bölüm/program ağacına bağlanır.

Doğru sıra:  import_yok_collector.py  →  import_ekdata.py
(YÖK önce çalışır çünkü hiyerarşinin omurgasını o kurar.)

TASARIM KURALLARI
-----------------
* Kaynakta olmayan değer UYDURULMAZ → NULL.
* Yıl granülerliği KORUNUR → her yıl ayrı kayıt.
* Köken (dosya, kaynak türü, satır anahtarı) KORUNUR.
* Yazım/seviye farkı yüzünden KOPYA BİRİM açılmaz (bkz. unit_matching).
* İki kaynak aynı alana farklı değer verirse mevcut değer KORUNUR ve
  çakışma `data_source_conflicts` tablosuna yazılır.
* İdempotent: aynı veriyle ikinci çalıştırma yeni kayıt eklemez.

ÇALIŞTIRMA
----------
    python import_ekdata.py                     # varsayılan klasör
    python import_ekdata.py --dir <yol>
    python import_ekdata.py --dry-run           # yazmadan rapor
===========================================================================
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import sys
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

BACKEND_DIR = Path(__file__).resolve().parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import select  # noqa: E402
from sqlalchemy.orm import Session  # noqa: E402

from app.database import SessionLocal, init_db  # noqa: E402
from app.models import (  # noqa: E402
    AcademicProgram,
    CurriculumCourse,
    DataSourceConflict,
    Department,
    Faculty,
    ProgramEnrollmentSnapshot,
    YksPlacementRecord,
)
from app.services.unit_matching import (  # noqa: E402
    UnitIndex,
    normalize_unit_name,
)
from app.services import student_count  # noqa: E402
from app.services.unit_types import classify_unit  # noqa: E402

VARSAYILAN_DIZIN = BACKEND_DIR.parent / "data" / "ekdata"


# ---------------------------------------------------------------------------
# Küçük yardımcılar
# ---------------------------------------------------------------------------


def _metin(v: Any) -> Optional[str]:
    """Boş hücreyi None'a çevirir. '' ile NULL aynı şey değildir."""
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _tam_sayi(v: Any) -> Optional[int]:
    s = _metin(v)
    if s is None:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def _ondalik(v: Any) -> Optional[Decimal]:
    s = _metin(v)
    if s is None:
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _akademik_yil(yil: int) -> str:
    """2025 → '2025-2026'. ÖSYM takvim yılı verir, sistem akademik yıl kullanır."""
    return f"{yil}-{yil + 1}"


def _parmak_izi(*parcalar: Any) -> str:
    ham = "|".join("" if p is None else str(p) for p in parcalar)
    return hashlib.sha256(ham.encode("utf-8")).hexdigest()


def _kanonik_deger(v: Any) -> str:
    """Değeri BİÇİMDEN bağımsız metne çevirir.

    CSV her hücreyi metin olarak verir ("2025", "10", ""), JSON ise
    yerel türlerle (2025, 10, null). Aynı veri iki dosyada farklı
    serileşir ve ikiz dosya tanınmaz. Bu fonksiyon ikisini aynı yazıya
    indirger; sayılar sayı, boşluk boşluk olarak karşılaştırılır.
    """
    if v is None:
        return ""
    s = str(v).strip()
    if s == "":
        return ""
    try:
        d = Decimal(s)
    except (InvalidOperation, ValueError):
        return s
    # 10 ile 10.0 ve 1.1 ile 1.10 aynı sayıdır.
    return format(d.normalize(), "f")


def _icerik_izi(tur: str, satirlar: List[Dict[str, Any]]) -> str:
    """Bir veri kümesinin BİÇİMDEN bağımsız içerik özeti.

    Aynı veriyi hem CSV hem JSON olarak veren kaynak klasörlerde, aynı
    satırları iki kez aktarmayı önler.
    """
    kanonik = [
        "|".join(f"{k}={_kanonik_deger(satir.get(k))}" for k in sorted(satir))
        for satir in satirlar
    ]
    kanonik.sort()
    return _parmak_izi(tur, "\n".join(kanonik))


def _kod_uret(ad: str, uzunluk: int = 12) -> str:
    """Yeni birim için okunabilir kod türetir."""
    sade = normalize_unit_name(ad)
    sozcukler = [w for w in sade.split() if w]
    if not sozcukler:
        return "KOD"
    if len(sozcukler) == 1:
        return sozcukler[0][:uzunluk]
    return "".join(w[:3] for w in sozcukler[:3])[:uzunluk]


def _tekil_kod(taban: str, kullanilan: set) -> str:
    kod, i = taban, 2
    while kod in kullanilan:
        kod = f"{taban[:10]}{i}"
        i += 1
    kullanilan.add(kod)
    return kod


# Bir hücrede bütün bir dönem tablosunun yapıştığını tanıyan sezgi:
# birden çok ders kodu veya aşırı uzunluk.
_DERS_KODU = re.compile(r"\b[A-ZÇĞİÖŞÜ]{2,6}\s?\d{3}\b")

#: Bölüm adına yapışmış öğretim dili eki — bölümün değil programın özelliği.
_DIL_EKI = re.compile(
    r"\s*\((?:İngilizce|Ingilizce|English|Türkçe|Turkce|Almanca|Fransızca)\)",
    re.IGNORECASE,
)


def _ad_guvenilir_mi(ad: str) -> bool:
    """PDF ayrıştırma artığı mı? (birleşmiş dönem tablosu)"""
    if not ad:
        return False
    return len(_DERS_KODU.findall(ad)) < 2 and len(ad) <= 80


# ---------------------------------------------------------------------------
# Dosya tanıma
# ---------------------------------------------------------------------------

#: (küme adı, alan imzası) — dosya ADINA değil, İÇERİĞİNE bakılır.
YKS_ALANLARI = {
    "academic_year", "faculty", "department", "program_name",
    "score_type", "scholarship_type", "quota", "placed_students",
}
OZET_ALANLARI = {
    "faculty", "department", "program_name", "observed_years",
    "estimated_4_cohort_placed_students",
}
MUFREDAT_ALANLARI = {"faculty", "department", "course_code", "course_name"}


def _tanı(yol: Path) -> Tuple[str, List[Dict[str, Any]], List[str]]:
    """Dosyayı okur ve (küme türü, satırlar, sütunlar) döndürür.

    Küme türü: "yks" | "yks_summary" | "mufredat" | "bilinmiyor"
    """
    uzanti = yol.suffix.lower()
    satirlar: List[Dict[str, Any]] = []
    sutunlar: List[str] = []

    if uzanti == ".csv":
        with yol.open(encoding="utf-8-sig", newline="") as f:
            okuyucu = csv.DictReader(f)
            sutunlar = list(okuyucu.fieldnames or [])
            satirlar = [dict(r) for r in okuyucu]
    elif uzanti == ".json":
        veri = json.loads(yol.read_text(encoding="utf-8"))
        if isinstance(veri, dict):
            # Tek anahtarlı sarmalayıcı olabilir.
            for deger in veri.values():
                if isinstance(deger, list):
                    veri = deger
                    break
        if isinstance(veri, list) and veri and isinstance(veri[0], dict):
            satirlar = veri
            sutunlar = list(veri[0].keys())
    elif uzanti in (".xlsx", ".xlsm"):
        try:
            import openpyxl
        except ImportError:  # pragma: no cover
            return ("bilinmiyor", [], [])
        wb = openpyxl.load_workbook(yol, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        basliklar = [str(h).strip() if h is not None else "" for h in next(it, [])]
        sutunlar = basliklar
        for satir in it:
            if all(h is None for h in satir):
                continue
            satirlar.append(dict(zip(basliklar, satir)))
        wb.close()
    else:
        return ("bilinmiyor", [], [])

    alan = set(sutunlar)
    if YKS_ALANLARI <= alan:
        return ("yks", satirlar, sutunlar)
    if OZET_ALANLARI <= alan:
        return ("yks_summary", satirlar, sutunlar)
    if MUFREDAT_ALANLARI <= alan:
        return ("mufredat", satirlar, sutunlar)
    return ("bilinmiyor", satirlar, sutunlar)


# ---------------------------------------------------------------------------
# Aktarım
# ---------------------------------------------------------------------------


class EkVeriAktarimi:
    """Tanınan kümeleri mevcut hiyerarşiye yazar."""

    def __init__(self, db: Session, dry_run: bool = False) -> None:
        self.db = db
        self.dry_run = dry_run
        self.dizin = UnitIndex(db)
        self.sayac: Counter = Counter()
        self.cakismalar: List[Dict[str, Any]] = []
        self.cozulemeyen: List[str] = []
        self.tanınmayan: List[str] = []
        self.notlar: List[str] = []
        self._bolum_kodlari = {
            c for (c,) in db.execute(select(Department.code))
        }
        self._program_kodlari = {
            c for (c,) in db.execute(select(AcademicProgram.code))
        }

    # -- çakışma korumalı yazma -------------------------------------
    def _guncelle(
        self,
        nesne: Any,
        alan: str,
        yeni: Any,
        kaynak: str,
        tablo: str,
        etiket: str,
    ) -> None:
        """Alanı yazar; DOLU ve FARKLI ise yazmaz, çakışmayı kaydeder.

        NULL üzerine yazmak çakışma değildir: farklı kaynaklar aynı
        varlığın FARKLI alanlarını doldurabilir ve bu beklenen durumdur.
        """
        if yeni is None:
            return
        mevcut = getattr(nesne, alan, None)
        if mevcut is None or mevcut == "":
            setattr(nesne, alan, yeni)
            return
        if self._ayni_deger(mevcut, yeni):
            return

        # Gerçek çakışma: iki kaynak farklı değer söylüyor.
        self.cakismalar.append({
            "table_name": tablo,
            "record_id": getattr(nesne, "id", None),
            "field_name": alan,
            "record_label": etiket,
            "existing_value": str(mevcut),
            "incoming_value": str(yeni),
            "incoming_source": kaynak,
        })
        self.sayac["cakisma"] += 1

    @staticmethod
    def _ayni_deger(mevcut: Any, yeni: Any) -> bool:
        """İki değer, SÜTUNUN saklayabildiği hassasiyette aynı mı?

        `minimum_admission_score` sütunu iki ondalık saklıyor (MoneyType).
        Kaynak 431.91149 veriyor, sütun 431.91 olarak geri veriyor. Bunu
        çakışma saymak SAHTE bir uyarı üretir ve gerçek çakışmaları
        gürültüye boğar. Karşılaştırma, mevcut değerin ondalık
        basamağına yuvarlanarak yapılır.

        Tam hassasiyet kaybolmuyor: ham değer `yks_placement_records`
        tablosunda 5 ondalıkla duruyor.
        """
        if isinstance(mevcut, Decimal) and isinstance(yeni, Decimal):
            # Mevcut değerin ondalık basamağına yuvarlayıp karşılaştır.
            us = mevcut.as_tuple().exponent
            if not isinstance(us, int):        # NaN / Infinity
                return mevcut == yeni
            return yeni.quantize(Decimal(1).scaleb(us)) == mevcut
        return str(mevcut) == str(yeni)

    # -- eksik gerçek birimi oluştur ---------------------------------
    def _bolum_ac(self, ad: str, fakulte: Faculty, kaynak: str) -> Department:
        # Bölüm adından ÖĞRETİM DİLİ ekini at: "(İngilizce)" bir programın
        # özelliğidir, bölümün adı değil. Aynı bölümün Türkçe ve İngilizce
        # programı olabilir; eki bölüm adına yazmak, ikinci program
        # geldiğinde ikinci bir bölüm açılmasına yol açardı.
        # Ham etiket kaybolmasın diye açıklamada saklanır.
        temiz = _DIL_EKI.sub("", ad).strip()
        kod = _tekil_kod(_kod_uret(temiz or ad), self._bolum_kodlari)
        d = Department(
            name=temiz or ad, code=kod, faculty_id=fakulte.id, is_active=True,
            description=f"Kaynak: {kaynak} · kaynaktaki etiket: {ad}",
        )
        self.db.add(d)
        self.db.flush()
        self.dizin.remember_department(d)
        self.sayac["bolum_eklendi"] += 1
        self.notlar.append(
            f"Yeni bölüm oluşturuldu: {ad} ({fakulte.name}) — kaynak {kaynak}"
        )
        return d

    def _program_ac(
        self, ad: str, bolum: Department, kaynak: str, seviye: str = "Bilinmiyor"
    ) -> AcademicProgram:
        kod = _tekil_kod(_kod_uret(ad), self._program_kodlari)
        p = AcademicProgram(
            name=ad, code=kod, department_id=bolum.id, degree_level=seviye,
            # Kontenjan/süre bu kaynaktan gelmiyorsa NULL kalır.
            duration_years=None, quota=None, is_active=True,
            description=f"Kaynak: {kaynak}",
        )
        self.db.add(p)
        self.db.flush()
        self.dizin.remember_program(p)
        self.sayac["program_eklendi"] += 1
        return p

    # ==================================================================
    # 1) YKS yerleştirme
    # ==================================================================
    def yks_aktar(self, satirlar: List[Dict[str, Any]], dosya: str) -> None:
        """ÖSYM satırlarını kaynak granülerliğinde yazar."""
        kume = "ankara_bilim_yks_4year"
        # (program_id, akademik_yıl) -> toplanacak değerler
        ozet: Dict[Tuple[int, str], Dict[str, Any]] = defaultdict(
            lambda: {
                "quota": None, "placed": None,
                "taban": [], "burslu_taban": [],
            }
        )

        for satir in satirlar:
            fak_ad = _metin(satir.get("faculty"))
            bol_ad = _metin(satir.get("department"))
            prog_ad = _metin(satir.get("program_name"))
            yil = _tam_sayi(satir.get("academic_year"))
            if not (fak_ad and bol_ad and prog_ad and yil):
                self.sayac["yks_eksik_anahtar"] += 1
                continue

            fakulte = self.dizin.find_faculty(fak_ad)
            if fakulte is None:
                self.cozulemeyen.append(f"[{dosya}] fakülte çözülemedi: {fak_ad}")
                self.sayac["yks_fakulte_cozulemedi"] += 1
                continue

            program = self._program_bul_veya_ac(bol_ad, fakulte, dosya)
            if program is None:
                self.sayac["yks_program_cozulemedi"] += 1
                continue

            anahtar = (
                yil, prog_ad,
                _metin(satir.get("score_type")) or "",
                _metin(satir.get("scholarship_type")) or "",
            )
            mevcut = self.db.execute(
                select(YksPlacementRecord).where(
                    YksPlacementRecord.placement_year == anahtar[0],
                    YksPlacementRecord.placement_program_name == anahtar[1],
                    YksPlacementRecord.score_type == anahtar[2],
                    YksPlacementRecord.scholarship_type == anahtar[3],
                )
            ).scalar_one_or_none()

            alanlar = dict(
                academic_program_id=program.id,
                academic_year=_akademik_yil(yil),
                placement_program_code=_metin(satir.get("program_code")),
                quota=_tam_sayi(satir.get("quota")),
                placed_students=_tam_sayi(satir.get("placed_students")),
                vacant_quota=_tam_sayi(satir.get("vacant_quota")),
                occupancy_rate=_ondalik(satir.get("occupancy_rate")),
                base_score=_ondalik(satir.get("base_score")),
                highest_score=_ondalik(satir.get("highest_score")),
                success_rank=_tam_sayi(satir.get("success_rank")),
                source_dataset=kume,
                source_file=dosya,
                source_row_key="|".join(str(a) for a in anahtar),
            )

            if mevcut is None:
                self.db.add(YksPlacementRecord(
                    placement_year=anahtar[0],
                    placement_program_name=anahtar[1],
                    score_type=anahtar[2],
                    scholarship_type=anahtar[3],
                    **alanlar,
                ))
                self.sayac["yks_eklendi"] += 1
            else:
                for k, v in alanlar.items():
                    setattr(mevcut, k, v)
                self.sayac["yks_guncellendi"] += 1

            # --- snapshot için topla ---
            kova = ozet[(program.id, _akademik_yil(yil))]
            for kaynak_alan, hedef in (("quota", "quota"),
                                       ("placed_students", "placed")):
                deger = _tam_sayi(satir.get(kaynak_alan))
                if deger is not None:
                    kova[hedef] = (kova[hedef] or 0) + deger
            taban = _ondalik(satir.get("base_score"))
            if taban is not None:
                kova["taban"].append(taban)
                if (_metin(satir.get("scholarship_type")) or "").lower() == "burslu":
                    kova["burslu_taban"].append(taban)

        self.db.flush()
        self._snapshot_yaz(ozet, kume, dosya)
        self._program_kontenjani_yaz(ozet, f"{kume} ({dosya})")

    def _program_bul_veya_ac(
        self, bol_ad: str, fakulte: Faculty, dosya: str
    ) -> Optional[AcademicProgram]:
        """Kaynağın "department" değerini KURUMSAL PROGRAMA bağlar.

        ÖSYM'nin "department"ı çoğu zaman bizim PROGRAMIMIZDIR
        (bkz. unit_matching). Sırayla:
          1. program adı eşleşiyor  → o program
          2. bölüm adı eşleşiyor    → o bölümün tek programı; yoksa aynı
                                      adla bir program açılır
          3. hiçbiri                → bölüm + program oluşturulur
        """
        eslesme = self.dizin.resolve(bol_ad, fakulte)

        if eslesme.level == "program":
            return self.db.get(AcademicProgram, eslesme.academic_program_id)

        if eslesme.level == "department":
            bolum = self.db.get(Department, eslesme.department_id)
            programlar = self.db.execute(
                select(AcademicProgram).where(
                    AcademicProgram.department_id == bolum.id
                )
            ).scalars().all()
            if len(programlar) == 1:
                return programlar[0]
            if not programlar:
                return self._program_ac(bol_ad, bolum, dosya)
            # Birden çok program varsa ad eşleşmesi denenir; bulunamazsa
            # KEYFÎ SEÇİM YAPILMAZ — yanlış programa veri yazmak,
            # veriyi hiç yazmamaktan kötüdür.
            hedef = normalize_unit_name(bol_ad)
            for p in programlar:
                if normalize_unit_name(p.name) == hedef:
                    return p
            self.cozulemeyen.append(
                f"[{dosya}] '{bol_ad}' bölümünde {len(programlar)} program var; "
                "hangisine ait olduğu kaynaktan anlaşılmıyor."
            )
            return None

        # Hiç yok: gerçek bir birim, oluşturulur.
        bolum = self._bolum_ac(bol_ad, fakulte, dosya)
        return self._program_ac(bol_ad, bolum, dosya)

    def _snapshot_yaz(
        self, ozet: Dict[Tuple[int, str], Dict[str, Any]], kume: str, dosya: str
    ) -> None:
        """Yerleştirme satırlarını (program, yıl) snapshot'ına TOPLAR.

        Mevcut API'ler, kapsam mantığı ve AI araçları snapshot okuyor;
        onları bozmadan gerçek veriyi görünür kılmanın yolu bu.

        Toplama kuralları:
          kontenjan / yerleşen : varyantların TOPLAMI (burslu + indirimli
                                 + ücretli aynı programın kontenjanıdır)
          taban puan           : varyantların EN DÜŞÜĞÜ — programa giriş
                                 eşiği budur
          tam burslu taban     : yalnızca "Burslu" varyantların en düşüğü
        """
        for (program_id, akademik_yil), kova in ozet.items():
            program = self.db.get(AcademicProgram, program_id)
            etiket = f"{program.name if program else program_id} · {akademik_yil}"

            mevcut = self.db.execute(
                select(ProgramEnrollmentSnapshot).where(
                    ProgramEnrollmentSnapshot.academic_program_id == program_id,
                    ProgramEnrollmentSnapshot.academic_year == akademik_yil,
                )
            ).scalar_one_or_none()

            kontenjan = kova["quota"]
            yerlesen = kova["placed"]
            taban = min(kova["taban"]) if kova["taban"] else None
            burslu = min(kova["burslu_taban"]) if kova["burslu_taban"] else None

            if mevcut is None:
                # quota NOT NULL; kaynakta yoksa kayıt açmak yanlış olur.
                if kontenjan is None:
                    self.sayac["snapshot_atlandi_kontenjan_yok"] += 1
                    continue
                self.db.add(ProgramEnrollmentSnapshot(
                    academic_program_id=program_id,
                    academic_year=akademik_yil,
                    quota=kontenjan,
                    enrolled_student_count=yerlesen or 0,
                    minimum_admission_score=taban,
                    full_scholarship_minimum_admission_score=burslu,
                ))
                self.sayac["snapshot_eklendi"] += 1
            else:
                kaynak = f"{kume} ({dosya})"
                self._guncelle(mevcut, "quota", kontenjan, kaynak,
                               "program_enrollment_snapshots", etiket)
                self._guncelle(mevcut, "enrolled_student_count", yerlesen, kaynak,
                               "program_enrollment_snapshots", etiket)
                self._guncelle(mevcut, "minimum_admission_score", taban, kaynak,
                               "program_enrollment_snapshots", etiket)
                self._guncelle(mevcut, "full_scholarship_minimum_admission_score",
                               burslu, kaynak, "program_enrollment_snapshots", etiket)
                self.sayac["snapshot_guncellendi"] += 1

    def _program_kontenjani_yaz(
        self, ozet: Dict[Tuple[int, str], Dict[str, Any]], kaynak: str
    ) -> None:
        """`academic_programs.quota` alanını EN GÜNCEL yıldan doldurur.

        Bu alan YÖK aktarımında NULL bırakılmıştı (toplayıcıda kontenjan
        yok). ÖSYM verisi gerçek kontenjanı getiriyor; belgelenmiş bir
        boşluk kapanıyor.
        """
        en_guncel: Dict[int, Tuple[str, Optional[int]]] = {}
        for (program_id, akademik_yil), kova in ozet.items():
            onceki = en_guncel.get(program_id)
            if onceki is None or akademik_yil > onceki[0]:
                en_guncel[program_id] = (akademik_yil, kova["quota"])

        for program_id, (yil, kontenjan) in en_guncel.items():
            if kontenjan is None:
                continue
            program = self.db.get(AcademicProgram, program_id)
            if program is None:
                continue
            self._guncelle(
                program, "quota", kontenjan, f"{kaynak} · {yil}",
                "academic_programs", program.name,
            )
            if program.quota == kontenjan:
                self.sayac["program_kontenjani_dolduruldu"] += 1

    # ==================================================================
    # 2) Müfredat
    # ==================================================================
    def mufredat_aktar(self, satirlar: List[Dict[str, Any]], dosya: str) -> None:
        kume = "ankara_bilim_mufredat"
        for satir in satirlar:
            fak_ad = _metin(satir.get("faculty"))
            bol_ad = _metin(satir.get("department"))
            ders_ad = _metin(satir.get("course_name"))
            if not (fak_ad and bol_ad and ders_ad):
                self.sayac["mufredat_eksik_anahtar"] += 1
                continue

            fakulte = self.dizin.find_faculty(fak_ad)
            if fakulte is None:
                self.cozulemeyen.append(f"[{dosya}] fakülte çözülemedi: {fak_ad}")
                self.sayac["mufredat_fakulte_cozulemedi"] += 1
                continue

            eslesme = self.dizin.resolve(bol_ad, fakulte)
            if eslesme.level == "program":
                bolum_id = eslesme.department_id
                program_id = eslesme.academic_program_id
            elif eslesme.level == "department":
                bolum_id, program_id = eslesme.department_id, None
            else:
                bolum = self._bolum_ac(bol_ad, fakulte, dosya)
                bolum_id, program_id = bolum.id, None

            kod = _metin(satir.get("course_code"))
            izi = _parmak_izi(
                kume, bol_ad, kod, ders_ad,
                _metin(satir.get("source_type")), _metin(satir.get("source")),
            )

            mevcut = self.db.execute(
                select(CurriculumCourse).where(
                    CurriculumCourse.source_fingerprint == izi
                )
            ).scalar_one_or_none()

            alanlar = dict(
                department_id=bolum_id,
                academic_program_id=program_id,
                course_code=kod,
                course_name=ders_ad,
                name_is_reliable=_ad_guvenilir_mi(ders_ad),
                source_unit_label=bol_ad,
                source_type=_metin(satir.get("source_type")) or "bilinmiyor",
                source_reference=_metin(satir.get("source")),
                source_dataset=kume,
                source_file=dosya,
            )
            if mevcut is None:
                self.db.add(CurriculumCourse(source_fingerprint=izi, **alanlar))
                self.sayac["mufredat_eklendi"] += 1
                if not alanlar["name_is_reliable"]:
                    self.sayac["mufredat_ad_guvenilmez"] += 1
            else:
                for k, v in alanlar.items():
                    setattr(mevcut, k, v)
                self.sayac["mufredat_guncellendi"] += 1
        self.db.flush()

    # ==================================================================
    # 3) Türetilmiş özet — bilinçli olarak SAKLANMAZ
    # ==================================================================
    def ozet_dogrula(self, satirlar: List[Dict[str, Any]], dosya: str) -> None:
        """Özet dosyayı DOĞRULAMA için okur, veritabanına yazmaz.

        `estimated_4_cohort_placed_students` adı gereği TAHMİNDİR ve bazı
        satırlarda boştur. Ölçülmüş bir alanmış gibi saklamak, uydurma
        veri yasağını çiğnerdi. Zaten yıllık satırlardan türetilebilir.

        Yine de sessizce atmıyoruz: yıl listesi ile 4year dosyasının
        tutarlı olup olmadığı kontrol edilir ve rapora yazılır.
        """
        tutarsiz = 0
        for satir in satirlar:
            yillar = (_metin(satir.get("observed_years")) or "").split("|")
            sayi = _tam_sayi(satir.get("years_with_placed_data"))
            if sayi is not None and sayi > len([y for y in yillar if y]):
                tutarsiz += 1
        self.sayac["ozet_satir_okundu"] += len(satirlar)
        self.notlar.append(
            f"[{dosya}] türetilmiş özet: {len(satirlar)} satır doğrulandı, "
            f"{tutarsiz} tutarsız; TAHMİN sütunu bilinçli olarak saklanmadı."
        )

    # ==================================================================
    # Çakışmaları kalıcı yaz
    # ==================================================================
    def cakismalari_yaz(self) -> None:
        for c in self.cakismalar:
            if c["record_id"] is None:
                continue
            var = self.db.execute(
                select(DataSourceConflict).where(
                    DataSourceConflict.table_name == c["table_name"],
                    DataSourceConflict.record_id == c["record_id"],
                    DataSourceConflict.field_name == c["field_name"],
                    DataSourceConflict.incoming_source == c["incoming_source"],
                )
            ).scalar_one_or_none()
            if var is not None:
                var.existing_value = c["existing_value"]
                var.incoming_value = c["incoming_value"]
                continue
            self.db.add(DataSourceConflict(
                resolution="kept_existing",
                note="Mevcut değer korundu; gelen kaynak farklı değer bildirdi.",
                **c,
            ))
        self.db.flush()


# ---------------------------------------------------------------------------
# Sürücü
# ---------------------------------------------------------------------------

BOS_KALAN_ALANLAR = [
    ("yks_placement_records.placement_program_code", "kaynakta 212/212 satırda boş"),
    ("yks_placement_records.vacant_quota", "kaynakta 212/212 satırda boş"),
    ("yks_placement_records.highest_score", "kaynakta 212/212 satırda boş"),
    ("curriculum_courses.course_code", "kaynakta 76/1205 satırda boş"),
    ("curriculum_courses (kredi/AKTS/yarıyıl)", "temiz satırlarda hiç yok — sütun açılmadı"),
]


def main() -> int:
    ap = argparse.ArgumentParser(description="data/ekdata gerçek verisini aktarır.")
    ap.add_argument("--dir", type=Path, default=VARSAYILAN_DIZIN,
                    help="ekdata klasörü")
    ap.add_argument("--dry-run", action="store_true",
                    help="Veritabanına yazmadan yalnızca rapor üret")
    args = ap.parse_args()

    if not args.dir.exists():
        raise SystemExit(f"Klasör bulunamadı: {args.dir}")

    print("=" * 68)
    print("EK GERÇEK VERİ (data/ekdata) → ÜRETİM VERİTABANI")
    print("=" * 68)
    print(f"Kaynak klasör : {args.dir}")
    if args.dry_run:
        print("MOD           : DRY-RUN (yazma yok)")

    init_db()
    db = SessionLocal()
    aktarim = EkVeriAktarimi(db, dry_run=args.dry_run)

    # `part2` ve `part3` BU BETİĞE AİT DEĞİLDİR; kendi aktarım
    # betikleri vardır (`import_yok_registry.py`, `import_part3.py`).
    # Burada taranırlarsa her çalıştırmada "tanınmayan şema"
    # gürültüsü üretir ve kullanıcıya aktarılmamış izlenimi verir.
    dosyalar = sorted(
        p for p in args.dir.rglob("*")
        if p.is_file()
        and not {"part2", "part3"} & set(p.relative_to(args.dir).parts)
    )
    print(f"\n[1/3] {len(dosyalar)} dosya tarandı")

    # JSON/CSV ikizini iki kez yazmamak için içerik özeti tutulur.
    gorulen_icerik: Dict[str, str] = {}

    try:
        for yol in dosyalar:
            goreli = str(yol.relative_to(args.dir))
            tur, satirlar, sutunlar = _tanı(yol)
            if tur == "bilinmiyor":
                aktarim.tanınmayan.append(f"{goreli} (sütunlar: {sutunlar[:6]})")
                print(f"      ATLANDI  {goreli}  — tanınmayan şema")
                continue

            izi = _icerik_izi(tur, satirlar)
            if izi in gorulen_icerik:
                aktarim.notlar.append(
                    f"{goreli}: içeriği {gorulen_icerik[izi]} ile BİREBİR AYNI; "
                    "ikinci kez aktarılmadı (aynı veri, farklı biçim)."
                )
                print(f"      IKIZ     {goreli}  — {gorulen_icerik[izi]} ile aynı")
                continue
            gorulen_icerik[izi] = goreli

            print(f"      {tur:12} {goreli}  ({len(satirlar)} satır)")
            if tur == "yks":
                aktarim.yks_aktar(satirlar, goreli)
            elif tur == "mufredat":
                aktarim.mufredat_aktar(satirlar, goreli)
            elif tur == "yks_summary":
                aktarim.ozet_dogrula(satirlar, goreli)

        print("\n[2/3] Müfredat kanonikleştiriliyor…")
        # Ham satırlara DOKUNULMAZ; uygulamanın okuduğu temiz katman
        # ham veriden yeniden türetilir.
        from app.services.curriculum_canonical import rebuild_canonical
        for anahtar, adet in rebuild_canonical(db).items():
            aktarim.sayac[f"mufredat_{anahtar}"] = adet

        print("      Resmî öğrenci sayıları hesaplanıyor…")
        # ÖSYM satırları yazıldıktan SONRA çalışır: sayı son ≤4 yerleştirme
        # yılının toplamıdır ve bütün satırlar yerinde olmalıdır.
        for anahtar, adet in student_count.refresh_stored_counts(db).items():
            aktarim.sayac[f"ogrenci_sayisi_{anahtar}"] += adet

        print("      Çakışmalar kaydediliyor…")
        aktarim.cakismalari_yaz()

        print("[3/3] Yazılıyor…")
        if args.dry_run:
            db.rollback()
            print("      DRY-RUN: değişiklikler geri alındı.")
        else:
            db.commit()
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()

    # ----------------------------------------------------------------
    print("\n" + "-" * 68)
    print("AKTARILAN KAYITLAR")
    print("-" * 68)
    for k in sorted(aktarim.sayac):
        print(f"  {k:38s} {aktarim.sayac[k]:>7,}")

    if aktarim.cakismalar:
        print("\n" + "-" * 68)
        print("ÇAKIŞMALAR (mevcut değer korundu)")
        print("-" * 68)
        for c in aktarim.cakismalar[:40]:
            print(f"  {c['table_name']}.{c['field_name']} · {c['record_label']}")
            print(f"      mevcut: {c['existing_value']}  ←  gelen: "
                  f"{c['incoming_value']}  ({c['incoming_source']})")
        if len(aktarim.cakismalar) > 40:
            print(f"  … {len(aktarim.cakismalar) - 40} çakışma daha "
                  "(data_source_conflicts tablosunda)")

    if aktarim.cozulemeyen:
        print("\n" + "-" * 68)
        print("ÇÖZÜLEMEYEN SATIRLAR (aktarılmadı)")
        print("-" * 68)
        for c in aktarim.cozulemeyen[:30]:
            print("  · " + c)

    if aktarim.tanınmayan:
        print("\n" + "-" * 68)
        print("TANINMAYAN DOSYALAR")
        print("-" * 68)
        for t in aktarim.tanınmayan:
            print("  · " + t)

    print("\n" + "-" * 68)
    print("KAYNAKTA OLMAYAN — BOŞ BIRAKILDI")
    print("-" * 68)
    for alan, aciklama in BOS_KALAN_ALANLAR:
        print(f"  {alan:48s} {aciklama}")

    if aktarim.notlar:
        print("\n" + "-" * 68)
        print("NOTLAR")
        print("-" * 68)
        for n in aktarim.notlar:
            print("  · " + n)

    print("\nTamamlandı. Betik idempotenttir; tekrar çalıştırmak kayıt eklemez.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
