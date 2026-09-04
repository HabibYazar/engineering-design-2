"""EKİP DERLEMESİ 2021-2025 — VARYANT DÜZEYİNDE AYRI AKTARIM.

NEDEN AYRI BİR KULVAR
---------------------
Ekipten gelen `2021_2025_.xlsx`, bir bölüm için TEK bir kaydı taşıyor:
ya tek bir YÖK program kodunu (Gazi'nin üç ayrı Bilgisayar Mühendisliği
programından yalnızca birini) ya da tek bir burs varyantını (yalnızca
"Burslu" kontenjanını).

Bizim mevcut serimiz (`source_dataset = "YÖK Atlas dataset 2025"`) ise bir
bölümün TÜM program kodlarının ve TÜM burs varyantlarının toplamıdır.

Bu iki tanım aynı kolona yazılırsa sahte bir çöküş üretir. Ölçüldü:
çakışan yıllarda kıyaslanabilen 44 kaydın 40'ında dosyanın değeri bizim
toplamımızın %60'ının altında; bir kısmı %4 seviyesinde. 2025'i olduğu
gibi eklersek her trend grafiği "kontenjan çöktü" der — oysa çöken şey
verinin kapsamı.

Bu yüzden bu betik AYRI bir `source_dataset` etiketiyle yazar. Mevcut
karşılaştırma servisleri zaten `source_dataset == SOURCE_DATASET` ile
filtreliyor; yani izolasyon koda güvenle değil, sorgunun kendisiyle
sağlanır. Yeni kulvarı görmek isteyen servis onu ADIYLA istemek
zorundadır.

NE AKTARILIR
------------
İKİ SAYFA, DÖRT METRİK:

* `KONTENJAN_YERLESEN_TUM` → quota, placed
  Yalnızca `Veri durumu` alanı "DOĞRULANDI" ile başlayan satırlar.
  "KONTROL BEKLİYOR", "KAYNAKTA DEĞER YOK" ve "temsilî" satırlar elenir.

* `ANA_KARSILASTIRMA` → success_rank, base_score
  Kanonik kulvarda sıralama yalnızca 2022-2024 için var; bu sayfa beş
  yılı birden taşıyor.

Her iki sayfada da 2021-2025'in tamamı alınır. Kanonik kulvara
DOKUNULMAZ; satırlar ayrı `source_dataset` etiketiyle yazılır.

UYDURMA İKİ SATIR ELENİR
------------------------
`ANA_KARSILASTIRMA`, kaynakta boş olan iki satırı başka satırların
puan/sıra değerleriyle doldurmuş (ABÜ Bilgisayar Mühendisliği 2025 ve
Çankaya Hukuk 2025). Eleme listesi elle yazılmaz: `BIZIM_WEB_VERILERI`
ve `AYNI_WEB_VERILERI` sayfalarında değeri boş olan satırlar okunarak
üretilir, böylece dosya düzeltilince eleme kendiliğinden kalkar.

HER SATIRA KAPSAM ETİKETİ YAZILIR
---------------------------------
`methodology` alanına, o program-üniversite çiftinin çakışan yıllardaki
kapsam sınamasının sonucu yazılır: TAM / EKSİK / SINANAMAZ. Böylece
ileride bu kulvarı kullanan herkes hangi satıra ne kadar
güvenebileceğini satırın kendisinden okur.

KULLANIM
--------
    python import_team_2021_2025.py                # kuru çalıştırma
    python import_team_2021_2025.py --uygula       # veritabanına yaz

Yazma kipi idempotenttir: kendi kulvarındaki satırları silip yeniden
kurar, başka hiçbir kulvara dokunmaz.
"""

from __future__ import annotations

import argparse
import glob
import hashlib
import json
import sys
from collections import Counter, defaultdict
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from sqlalchemy import delete, select

sys.path.insert(0, str(Path(__file__).resolve().parent))

from app.database import SessionLocal  # noqa: E402
from app.models import YokAtlasBenchmarkMetric  # noqa: E402
from app.services.program_equivalence import canonical_program_key  # noqa: E402
from app.services.yok_atlas_comparison_service import SOURCE_DATASET  # noqa: E402

#: Bu betiğin yazdığı kulvarın adı. Mevcut `SOURCE_DATASET`ten FARKLIDIR
#: ve bu fark kasıtlıdır — izolasyonun tamamı buna dayanır.
EKIP_DATASET = "Ekip derlemesi 2021-2025 (varyant düzeyi)"
EKIP_SOURCE_FILE = "2021_2025_.xlsx"
#: `BIZIM_WEB_VERILERI` sayfasında üniversite kolonu yoktur; satırlar
#: zaten kendi kurumumuza aittir.
HOME_ADI = "ANKARA BİLİM ÜNİVERSİTESİ"

SAYFA = "KONTENJAN_YERLESEN_TUM"
#: DOSYANIN TAMAMI ALINIR — yalnızca 2021/2025 değil.
#: ---------------------------------------------------------------------
#: İlk sürüm 2022-2024'ü eliyordu; gerekçe "kanonik kulvarımız daha
#: eksiksiz" idi ve o gerekçe kanonik kulvar için hâlâ doğru. Ama bu
#: kulvar kanonik kulvarın yerine geçmiyor, YANINDA duruyor: burada
#: dosyanın kendi iç tutarlılığı önemli. 2022-2024'ü dışarıda bırakmak,
#: aynı üniversite-bölüm çiftinin ekip kulvarındaki serisini ortasından
#: kesiyordu.
#:
#: Kanonik kulvara hâlâ DOKUNULMAZ; bu satırlar ayrı `source_dataset`
#: etiketiyle yazılır ve yalnızca onu adıyla isteyen sorgu görür.
HEDEF_YILLAR = (2021, 2022, 2023, 2024, 2025)

#: `Veri durumu` metninin sonundaki burs varyantı ipucu → kolon değeri.
#: Sıra önemli: "Burslu program" önce eşleşmeli, yoksa "Burslu" onu yutar.
BURS_IPUCU: Tuple[Tuple[str, Optional[str]], ...] = (
    ("%50 İndirimli", "%50 İndirimli"),
    ("Burslu program", "Burslu"),
    ("Burslu", "Burslu"),
    ("Ücretli", "Ücretli"),
    ("Devlet", None),          # devlet üniversitesinde burs varyantı yok
)

METRIK_KOLON = (("Kontenjan", "quota"), ("Yerleşen", "placed"))

#: PUAN VE SIRALAMA AYRI BİR SAYFADAN GELİR.
#: ---------------------------------------------------------------------
#: `KONTENJAN_YERLESEN_TUM` yalnızca adet taşır. Taban puan ve başarı
#: sırası `ANA_KARSILASTIRMA` sayfasındadır ve orada 2021-2025'in beşi de
#: bulunur — kanonik kulvarda ise sıralama yalnızca 2022-2024 için var.
SIRA_SAYFA = "ANA_KARSILASTIRMA"
SIRA_BOLUM_KOLON = "Ankara Bilim Bölümü"
SIRA_METRIK = (("Başarı Sırası", "success_rank"), ("Taban Puan", "base_score"))

#: Kaynak sayfalarında DEĞERİ BOŞ olan satırlar.
#: ---------------------------------------------------------------------
#: `ANA_KARSILASTIRMA` bu iki satırı boş bırakmak yerine başka satırların
#: değerleriyle doldurmuş. Kanıtı: aynı (taban puan, başarı sırası) çifti
#: birbiriyle ilgisiz iki satırda görünüyor ve biri SAY, diğeri EA puan
#: türünde — farklı puan türlerinde aynı sıra imkânsızdır.
#:
#: Liste elle yazılmadı; `BIZIM_WEB_VERILERI` ve `AYNI_WEB_VERILERI`
#: sayfalarında puanı boş olan satırlar okunarak üretilir. Dosya
#: düzeltilirse bu eleme kendiliğinden devre dışı kalır.
KAYNAK_SAYFALAR = (
    ("BIZIM_WEB_VERILERI", "Bölüm"),
    ("AYNI_WEB_VERILERI", "Bölüm"),
)


def _sayi(v) -> Optional[float]:
    try:
        return float(str(v).replace(",", "."))
    except (TypeError, ValueError):
        return None


def _burs(durum: str) -> Optional[str]:
    for ipucu, deger in BURS_IPUCU:
        if ipucu in durum:
            return deger
    return None


def _program_kodu(pkey: str, uni: str, burs: Optional[str], yil: int) -> str:
    """Kaynakta YÖK program kodu yok; kararlı bir vekil üretilir.

    Vekil, satırın kimliğinden türetilir. Aynı satır ikinci kez
    aktarıldığında aynı kodu üretir; tablodaki tekillik kısıtı
    (source_file, source_program_code, source_year, metric) bu sayede
    mükerrer kaydı veritabanı düzeyinde engeller.
    """
    ham = f"{pkey}|{uni}|{burs or '-'}|{yil}"
    return "EKIP-" + hashlib.sha1(ham.encode("utf-8")).hexdigest()[:12]


# ---------------------------------------------------------------------------
# Dosyayı oku
# ---------------------------------------------------------------------------
def _sayfa(wb, ad: str) -> List[dict]:
    ws = wb[ad]
    basliklar = [c.value for c in ws[1]]
    return [
        dict(zip(basliklar, satir))
        for satir in ws.iter_rows(min_row=2, values_only=True)
        if any(satir)
    ]


def dosyayi_oku(yol: Path) -> List[dict]:
    import openpyxl

    return _sayfa(openpyxl.load_workbook(yol, data_only=True), SAYFA)


def sira_oku(yol: Path) -> Tuple[List[dict], set]:
    """Puan/sıra satırları + kaynakta BOŞ olduğu için güvenilmeyecek anahtarlar."""
    import openpyxl

    wb = openpyxl.load_workbook(yol, data_only=True)
    supheli = set()
    for sayfa, bolum_kolon in KAYNAK_SAYFALAR:
        if sayfa not in wb.sheetnames:
            continue
        for r in _sayfa(wb, sayfa):
            bos = _sayi(r.get("Taban Puan")) is None or _sayi(r.get("Başarı Sırası")) is None
            if not bos:
                continue
            supheli.add((
                canonical_program_key(str(r.get(bolum_kolon) or "")),
                str(r.get("Üniversite") or HOME_ADI).strip().upper(),
                str(r.get("Yıl") or ""),
            ))
    return _sayfa(wb, SIRA_SAYFA), supheli


# ---------------------------------------------------------------------------
# Kapsam sınaması — dosya bu çifti çakışan yıllarda TAM kapsıyor mu?
# ---------------------------------------------------------------------------
def kapsam_haritasi(db, satirlar: List[dict]) -> Dict[Tuple[str, str], str]:
    """(program, üniversite) → "TAM" | "EKSİK" | "SINANAMAZ".

    Sınama şu: dosyanın 2022/2023/2024 kontenjanı, bizim aynı yıldaki TAM
    toplamımıza eşit mi? Eşitse dosya o çifti eksiksiz kapsıyor demektir
    ve aynı çiftin 2021/2025 satırına da güvenilebilir. Değilse dosya o
    çiftte yalnızca bir parçayı taşıyor.
    """
    bizim: Dict[Tuple[str, str, int], float] = defaultdict(float)
    for pkey, uni, yil, deger in db.execute(
        select(
            YokAtlasBenchmarkMetric.canonical_program_key,
            YokAtlasBenchmarkMetric.university_name,
            YokAtlasBenchmarkMetric.source_year,
            YokAtlasBenchmarkMetric.value,
        ).where(
            YokAtlasBenchmarkMetric.metric == "quota",
            YokAtlasBenchmarkMetric.source_dataset == SOURCE_DATASET,
        )
    ):
        bizim[(pkey, str(uni).strip().upper(), int(yil))] += float(deger or 0)

    sonuc: Dict[Tuple[str, str], List[bool]] = defaultdict(list)
    for r in satirlar:
        yil = int(r["Yıl"])
        if yil not in (2022, 2023, 2024):
            continue
        cift = (canonical_program_key(str(r["Bölüm"])), str(r["Üniversite"]).strip().upper())
        deger = _sayi(r.get("Kontenjan"))
        anahtar = (cift[0], cift[1], yil)
        if deger is None or anahtar not in bizim:
            continue
        sonuc[cift].append(abs(bizim[anahtar] - deger) < 0.5)

    return {
        cift: ("TAM" if all(bayraklar) else "EKSİK")
        for cift, bayraklar in sonuc.items()
        if bayraklar
    }


# ---------------------------------------------------------------------------
# Üniversite üstverisi — şehir/tür mevcut satırlardan devralınır
# ---------------------------------------------------------------------------
def uni_ustverisi(db) -> Dict[str, dict]:
    out: Dict[str, dict] = {}
    for uni, sehir, tur, ukod in db.execute(
        select(
            YokAtlasBenchmarkMetric.university_name,
            YokAtlasBenchmarkMetric.city,
            YokAtlasBenchmarkMetric.university_type,
            YokAtlasBenchmarkMetric.source_university_code,
        ).distinct()
    ):
        out.setdefault(str(uni).strip().upper(),
                       {"city": sehir, "type": tur, "code": ukod,
                        "name": uni})
    return out


def fakulte_haritasi(db) -> Dict[Tuple[str, str], Tuple[str, Optional[str]]]:
    """(program, üniversite) → en sık görülen (fakülte adı, kanonik anahtar)."""
    sayac: Dict[Tuple[str, str], Counter] = defaultdict(Counter)
    for pkey, uni, fak, fkey in db.execute(
        select(
            YokAtlasBenchmarkMetric.canonical_program_key,
            YokAtlasBenchmarkMetric.university_name,
            YokAtlasBenchmarkMetric.faculty_name,
            YokAtlasBenchmarkMetric.canonical_faculty_key,
        )
    ):
        sayac[(pkey, str(uni).strip().upper())][(fak, fkey)] += 1
    return {k: c.most_common(1)[0][0] for k, c in sayac.items()}


# ---------------------------------------------------------------------------
# Aktarım
# ---------------------------------------------------------------------------
def calistir(yol: Path, uygula: bool) -> int:
    satirlar = dosyayi_oku(yol)
    db = SessionLocal()
    try:
        kapsam = kapsam_haritasi(db, satirlar)
        ustveri = uni_ustverisi(db)
        fakulteler = fakulte_haritasi(db)

        eklenecek: List[YokAtlasBenchmarkMetric] = []
        elenen: Counter = Counter()
        kapsam_dagilim: Counter = Counter()

        for r in satirlar:
            yil = int(r["Yıl"])
            if yil not in HEDEF_YILLAR:
                elenen["yıl 2022-2024 (bizimki daha eksiksiz)"] += 1
                continue

            durum = str(r.get("Veri durumu") or "").strip()
            if not durum.startswith("DOĞRULANDI"):
                elenen[f"doğrulanmamış: {durum[:38]}"] += 1
                continue

            pkey = canonical_program_key(str(r["Bölüm"]))
            uni_ham = str(r["Üniversite"]).strip().upper()
            if not pkey:
                elenen["bölüm adı kanonikleştirilemedi"] += 1
                continue
            if uni_ham not in ustveri:
                elenen[f"üniversite eşleşmedi: {uni_ham[:30]}"] += 1
                continue

            burs = _burs(durum)
            cift = (pkey, uni_ham)
            kapsam_etiketi = kapsam.get(cift, "SINANAMAZ")
            kapsam_dagilim[kapsam_etiketi] += 1

            u = ustveri[uni_ham]
            fak, fkey = fakulteler.get(cift, ("(kaynakta belirtilmemiş)", None))
            kod = _program_kodu(pkey, uni_ham, burs, yil)

            for kolon, metrik in METRIK_KOLON:
                deger = _sayi(r.get(kolon))
                if deger is None:
                    elenen[f"{kolon} boş"] += 1
                    continue
                eklenecek.append(
                    YokAtlasBenchmarkMetric(
                        university_name=u["name"],
                        faculty_name=fak,
                        canonical_faculty_key=fkey,
                        program_name=str(r["Bölüm"]),
                        canonical_program_key=pkey,
                        city=u["city"],
                        university_type=u["type"],
                        program_language=None,
                        scholarship_type=burs,
                        source_description=None,
                        source_year=yil,
                        academic_year=f"{yil}-{yil + 1}",
                        metric=metrik,
                        value=Decimal(str(deger)),
                        source_raw_value=str(r.get(kolon)),
                        unit="kişi",
                        source_dataset=EKIP_DATASET,
                        source_file=EKIP_SOURCE_FILE,
                        source_program_code=kod,
                        source_university_code=u["code"],
                        source_row_identity=json.dumps(
                            {
                                "program": str(r["Bölüm"]),
                                "university": u["name"],
                                "year": yil,
                                "scholarship": burs,
                                "status": durum,
                                "source_url": r.get("Kaynak"),
                            },
                            ensure_ascii=False,
                        ),
                        derived=False,
                        methodology=(
                            f"Ekip derlemesi; kaynak durumu '{durum}'. "
                            f"KAPSAM={kapsam_etiketi}. Bu satır VARYANT "
                            f"DÜZEYİNDEDİR: bir bölümün tüm program kodlarının "
                            f"ve burs varyantlarının toplamı DEĞİLDİR. "
                            f"'{SOURCE_DATASET}' kulvarındaki yıl serisiyle "
                            f"toplanamaz veya doğrudan kıyaslanamaz."
                        ),
                    )
                )

        # ------------------------------------------------------------------
        # PUAN VE BAŞARI SIRASI
        # ------------------------------------------------------------------
        sira_satirlar, supheli = sira_oku(yol)
        for r in sira_satirlar:
            yil = int(str(r.get("Yıl") or 0) or 0)
            if yil not in HEDEF_YILLAR:
                continue
            pkey = canonical_program_key(str(r.get(SIRA_BOLUM_KOLON) or ""))
            uni_ham = str(r.get("Üniversite") or "").strip().upper()
            if not pkey or uni_ham not in ustveri:
                elenen["sıra: bölüm/üniversite eşleşmedi"] += 1
                continue
            if (pkey, uni_ham, str(yil)) in supheli:
                # Kaynak sayfada boş olan, `ANA_KARSILASTIRMA`'da başka bir
                # satırdan kopyalanmış değer. Aktarılırsa uydurma sayı
                # kalıcılaşırdı.
                elenen["sıra: kaynakta boş — kopyalanmış değer"] += 1
                continue

            u = ustveri[uni_ham]
            fak, fkey = fakulteler.get((pkey, uni_ham), ("(kaynakta belirtilmemiş)", None))
            puan_turu = str(r.get("Puan Türü") or "").strip() or None
            # Sıra satırının kimliği burs varyantı taşımaz; ayrı bir uzay.
            kod = _program_kodu(pkey, uni_ham, f"SIRA-{puan_turu or '-'}", yil)

            for kolon, metrik in SIRA_METRIK:
                deger = _sayi(r.get(kolon))
                if deger is None:
                    elenen[f"sıra: {kolon} boş"] += 1
                    continue
                eklenecek.append(
                    YokAtlasBenchmarkMetric(
                        university_name=u["name"],
                        faculty_name=fak,
                        canonical_faculty_key=fkey,
                        program_name=str(r.get(SIRA_BOLUM_KOLON)),
                        canonical_program_key=pkey,
                        city=u["city"],
                        university_type=u["type"],
                        program_language=None,
                        scholarship_type=None,
                        # Puan türü sıralamanın hangi evrende ölçüldüğünü
                        # söyler; SAY sırası ile EA sırası kıyaslanamaz.
                        source_description=puan_turu,
                        source_year=yil,
                        academic_year=f"{yil}-{yil + 1}",
                        metric=metrik,
                        value=Decimal(str(deger)),
                        source_raw_value=str(r.get(kolon)),
                        unit="sıra" if metrik == "success_rank" else "puan",
                        source_dataset=EKIP_DATASET,
                        source_file=EKIP_SOURCE_FILE,
                        source_program_code=kod,
                        source_university_code=u["code"],
                        source_row_identity=json.dumps(
                            {"program": str(r.get(SIRA_BOLUM_KOLON)),
                             "university": u["name"], "year": yil,
                             "score_type": puan_turu,
                             "match": r.get("Eşleşme"),
                             "source_url": r.get("Kaynak")},
                            ensure_ascii=False),
                        derived=False,
                        methodology=(
                            f"Ekip derlemesi ({SIRA_SAYFA}); puan türü "
                            f"{puan_turu or 'belirtilmemiş'}. Başarı sırası "
                            f"puan türüne göre ayrı bir sıralamadır; farklı "
                            f"türler karşılaştırılamaz."
                        ),
                    )
                )

        mevcut = db.execute(
            select(YokAtlasBenchmarkMetric.id).where(
                YokAtlasBenchmarkMetric.source_dataset == EKIP_DATASET
            )
        ).scalars().all()

        print(f"Dosya           : {yol.name}  ({len(satirlar)} satır)")
        print(f"Hedef kulvar    : {EKIP_DATASET}")
        print(f"Kulvarda mevcut : {len(mevcut)} satır (yazma kipinde silinip yeniden kurulur)")
        print(f"\nEKLENECEK metrik satırı: {len(eklenecek)}")
        print(f"   kapsam dağılımı (kayıt bazında): {dict(kapsam_dagilim)}")
        print("\nELENEN satırlar:")
        for sebep, adet in elenen.most_common():
            print(f"   {adet:>4}  {sebep}")

        if not uygula:
            print("\nKURU ÇALIŞTIRMA — veritabanına yazılmadı. Yazmak için: --uygula")
            return 0

        db.execute(
            delete(YokAtlasBenchmarkMetric).where(
                YokAtlasBenchmarkMetric.source_dataset == EKIP_DATASET
            )
        )
        db.add_all(eklenecek)
        db.commit()
        print(f"\nYAZILDI: {len(eklenecek)} satır → {EKIP_DATASET}")
        return 0
    finally:
        db.close()


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--dosya", default=None, help="xlsx yolu")
    p.add_argument("--uygula", action="store_true", help="veritabanına yaz")
    a = p.parse_args()

    if a.dosya:
        yol = Path(a.dosya)
    else:
        adaylar = sorted(Path(__file__).parent.glob("sample_data/*2021_2025*.xlsx"))
        if not adaylar:
            print("Dosya bulunamadı. --dosya ile yol verin.")
            return 2
        yol = adaylar[0]
    if not yol.exists():
        print(f"Dosya yok: {yol}")
        return 2
    return calistir(yol, a.uygula)


if __name__ == "__main__":
    sys.exit(main())
