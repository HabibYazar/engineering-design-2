"""DERSLİK KULLANIM HARİTASI — türetilmiş veri kümesi üreticisi.

NE YAPAR
--------
`data/infrastructure/raw/` altındaki DEĞİŞTİRİLMEYEN kaynaklardan
(3 kat planı PDF + derslik programı XLSX) deterministik, salt-okunur bir
normalize veri kümesi üretir:

    data/infrastructure/derived/classroom_usage.json    → kat/oda/slot/program
    data/infrastructure/derived/room_match_report.json  → eşleşme + anomali

NE YAPMAZ
---------
* Ham dosyaları OKUMAKTAN başka bir şey yapmaz (yazma/silme/yeniden
  adlandırma YOK).
* Veritabanına DOKUNMAZ. Import değildir, seed değildir.
* Eksik veriyi TAMAMLAMAZ. Kapasitesi yazmayan oda `null` kalır;
  programı olmayan oda `has_schedule: false` olur. Uydurma yoktur.

ÇALIŞTIRMA
----------
    python build_infrastructure_map.py

Çıktı idempotenttir: aynı girdi → aynı çıktı (sözlükler sıralı yazılır).
"""

from __future__ import annotations

import json
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl

KOK = Path(__file__).resolve().parent.parent          # integration/
HAM = KOK / "data" / "infrastructure" / "raw"
TURETILMIS = KOK / "data" / "infrastructure" / "derived"

XLSX = HAM / "Derslik Planı 2025-26 Bahar (1).xlsx"

#: KAT PLANI PDF'LERİ — DOSYA ADI ↔ İÇ BAŞLIK UYUŞMAZLIĞI
#: ----------------------------------------------------------------------
#: Dosya adları "0/1/2. Kat" derken PDF'lerin İÇİNDEKİ başlıklar bodrum
#: numaralandırması kullanıyor ve kendi içinde tutarsız:
#:
#:   0. Kat.pdf → "0. BODRUM KAT (D) PLANI (-13.20 KOTU)"
#:   1. Kat.pdf → "2. BODRUM KAT PLANI (-9.20 KOTU)"
#:   2. Kat.pdf → "1. BODRUM KAT PLANI (-5.20 KOTU)"
#:
#: Hangisi doğru? KOT (yükseklik) kararı verir; mimarın bodrum
#: numaralandırması değil:
#:     -13.20 < -9.20 < -5.20   →   0. Kat < 1. Kat < 2. Kat
#: Yani DOSYA ADI SIRASI kot sırasıyla ÖRTÜŞÜR.
#:
#: Bağımsız ikinci kanıt: 2. Kat.pdf içinde "AMFİ-3/4/5" etiketleri var;
#: XLSX'te de AMFİ'ler KAT 2'de listeli. 0. Kat.pdf'te "AMFİ" geçiyor;
#: XLSX'te LECTURE HALL 5/6/7 KAT 0'da. İki bağımsız kanıt aynı yöne
#: işaret ettiği için eşleme TAHMİN DEĞİL, doğrulanmıştır.
KAT_KAYNAKLARI = [
    {"floor": 0, "pdf": "0. Kat.pdf",
     "internal_title": "0. BODRUM KAT (D) PLANI (-13.20 KOTU)",
     "elevation_m": -13.20},
    {"floor": 1, "pdf": "1. Kat.pdf",
     "internal_title": "2. BODRUM KAT PLANI (-9.20 KOTU)",
     "elevation_m": -9.20},
    {"floor": 2, "pdf": "2. Kat.pdf",
     "internal_title": "1. BODRUM KAT PLANI (-5.20 KOTU)",
     "elevation_m": -5.20},
]

#: KAT_PLANI sayfasında her katın sütun bloğu (0 tabanlı sütun indeksi).
KAT_SUTUNLARI = {0: 0, 1: 6, 2: 12}

GUNLER = ["Pazartesi", "Salı", "Çarşamba", "Perşembe", "Cuma"]

#: Program/veri sayfası olmayan yardımcı sayfalar.
ODA_DISI_SAYFALAR = {"KAT_PLANI", "DERS_PLANI", "Eski_KAT_PLANI", "Sayfa1"}

#: GEÇERLİ ODA KODU BİÇİMLERİ — KAT_PLANI sütununda oda kodlarının yanı
#: sıra açıklama/efsane satırları da var ("Ders Planlama Esasları",
#: "1. Sabah 09.00 …", birim adları "İTBF"/"MMF" gibi). Bunlar oda
#: DEĞİLDİR. "Boş olmayan her hücre odadır" varsayımı envanteri 95'e
#: şişiriyordu. Kabul edilen biçimler kaynakta GÖZLENENLERDİR:
#:     C 003 / C071 / L 008 / L 143
#:     AMFİ1 / AMFİ 5 / Amfi-4
#:     LECTURE HALL 5
#: Bu listeye uymayan bir kod sessizce ATILMAZ; raporda görünür.
ODA_KODU_BICIMLERI = (
    re.compile(r"^[CL]\s?\d{2,3}$", re.IGNORECASE),
    re.compile(r"^AMF[İI]\s?-?\s?\d{1,2}$", re.IGNORECASE),
    re.compile(r"^LECTURE\s+HALL\s+\d{1,2}$", re.IGNORECASE),
)


def oda_kodu_mu(kod: str) -> bool:
    return any(k.match(kod.strip()) for k in ODA_KODU_BICIMLERI)

_TR = str.maketrans("ÇĞİIÖŞÜçğıioöşü", "CGIIOSUcgiioosu")


def oda_anahtari(ad: object) -> str:
    """Oda adını karşılaştırılabilir tek bir anahtara indirger.

    "C 220" / "C-220" / "c220"      → "C220"
    "AMFİ4" / "Amfi-4" / "AMFI 4"   → "AMFI4"
    "LECTURE HALL 5"                → "LECTUREHALL5"

    Türkçe'ye duyarlıdır: `upper()` "i"yi "I" yapar, oysa doğru karşılık
    "İ"dir; bu yüzden harfler önce elle eşlenir. Boşluk, tire, nokta ve
    alt çizgi ATILIR — bunlar yazım tercihidir, kimlik değildir.

    DİKKAT: Bu fonksiyon iki FARKLI odayı aynı anahtara indirgerse
    çağıran taraf hata vermelidir (bkz. `_anahtar_cakismasi_denetle`).
    """
    s = str(ad if ad is not None else "").strip()
    s = s.replace("İ", "I").replace("ı", "i")
    s = s.upper().translate(_TR)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[\s\-_.]+", "", s)


def _sayi(v: object) -> Optional[int]:
    """Kapasite hücresi → tam sayı. Boş/geçersizse None — 0 DEĞİL.

    0 yazmak "kapasite sıfır" demektir; boş hücre ise "ölçülmemiş"tir.
    İkisini karıştırmak doluluk oranını sessizce bozardı.
    """
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        n = int(float(str(v).strip().replace(",", ".")))
    except (TypeError, ValueError):
        return None
    return n


def _metin(v: object) -> Optional[str]:
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v)).strip()
    return s or None


def _dolu_mu(hucre: object) -> Optional[str]:
    """Program hücresi → ders adı (dolu) veya None (boş).

    Kaynakta boş slot "0" ile gösterilir. "0" bir ders DEĞİLDİR.
    """
    s = _metin(hucre)
    if s is None or s in {"0", "0.0", "-"}:
        return None
    return s


def kat_planini_oku(wb) -> tuple[List[dict], List[dict]]:
    """KAT_PLANI sayfasından oda envanteri. Kapasiteler buradan gelir.

    Dönüş: (odalar, oda_sayilmayanlar). İkincisi raporlanır ki bir oda
    biçim listesine uymadığı için sessizce kaybolmasın.
    """
    ws = wb["KAT_PLANI"]
    satirlar = list(ws.iter_rows(values_only=True))
    odalar: List[dict] = []
    elenenler: List[dict] = []
    for kat, sut in KAT_SUTUNLARI.items():
        for r in satirlar[3:]:                       # başlık satırlarını atla
            kod = _metin(r[sut] if sut < len(r) else None)
            if not kod:
                continue
            if not oda_kodu_mu(kod):
                elenenler.append({"value": kod[:70], "column_floor": kat,
                                  "reason": "oda_kodu_bicimine_uymuyor"})
                continue
            odalar.append({
                "room_code": kod,
                "room_key": oda_anahtari(kod),
                "floor": kat,
                "name": _metin(r[sut + 1]) if sut + 1 < len(r) else None,
                "class_capacity": _sayi(r[sut + 2]) if sut + 2 < len(r) else None,
                "student_capacity": _sayi(r[sut + 3]) if sut + 3 < len(r) else None,
                "owner_unit": _metin(r[sut + 4]) if sut + 4 < len(r) else None,
            })
    return odalar, elenenler


def program_sayfalarini_oku(wb) -> Dict[str, dict]:
    """Her derslik sayfasından gün × slot programı.

    Sayfalar iki farklı saat YAZIMI kullanıyor ("09:10-10:00" ve
    "09.00 - 09.50"). Slot SAYISI her sayfada 12'dir, bu yüzden kanonik
    kimlik SIRA NUMARASIDIR; sayfanın kendi etiketi ayrıca saklanır ve
    ipucunda gösterilir. Etiketlerden birini diğerine çevirmek, kaynakta
    olmayan bir saat iddiası olurdu.
    """
    sonuc: Dict[str, dict] = {}
    for sayfa in wb.sheetnames:
        if sayfa in ODA_DISI_SAYFALAR:
            continue
        ws = wb[sayfa]
        # Başlık satırını bul: "Saat/Gün" veya ilk sütunu saat olan satır.
        bas = None
        for r in range(1, min(ws.max_row, 6) + 1):
            hucre = _metin(ws.cell(r, 1).value) or ""
            if hucre.lower().startswith("saat") or re.match(r"^\d{2}[.:]\d{2}", hucre):
                bas = r
                break
        if bas is None:
            continue
        # "Saat/Gün" başlıksa slotlar bir alt satırdan başlar.
        ilk = bas + 1 if (_metin(ws.cell(bas, 1).value) or "").lower().startswith("saat") else bas

        slot_etiketleri: List[str] = []
        izgara: Dict[str, List[Optional[str]]] = {g: [] for g in GUNLER}
        for r in range(ilk, ws.max_row + 1):
            etiket = _metin(ws.cell(r, 1).value)
            if not etiket or not re.match(r"^\d{1,2}[.: ]", etiket):
                continue
            slot_etiketleri.append(etiket)
            for i, gun in enumerate(GUNLER, start=2):
                izgara[gun].append(_dolu_mu(ws.cell(r, i).value))
        if not slot_etiketleri:
            continue
        sonuc[oda_anahtari(sayfa)] = {
            "sheet_name": sayfa,
            "slot_labels": slot_etiketleri,
            "grid": izgara,
        }
    return sonuc


def _anahtar_cakismasi_denetle(odalar: List[dict]) -> List[dict]:
    """Farklı oda kodları aynı anahtara düşerse RAPORLA.

    Normalize etmek kimlikleri birleştirmemelidir. Aynı odanın iki
    yazımı ("C 071"/"C071") birleşmelidir; iki AYRI oda birleşmemelidir.
    Ayrım: aynı anahtar + aynı kat + aynı kapasite → aynı oda.
    """
    gruplar: Dict[str, List[dict]] = {}
    for o in odalar:
        gruplar.setdefault(o["room_key"], []).append(o)
    catismalar = []
    for anahtar, grup in gruplar.items():
        if len(grup) < 2:
            continue
        kimlik = {(g["floor"], g["class_capacity"], g["student_capacity"]) for g in grup}
        catismalar.append({
            "room_key": anahtar,
            "variants": [g["room_code"] for g in grup],
            "floors": sorted({g["floor"] for g in grup}),
            "same_room": len(kimlik) == 1,
        })
    return catismalar


def uret() -> dict:
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    odalar, elenenler = kat_planini_oku(wb)
    programlar = program_sayfalarini_oku(wb)
    catismalar = _anahtar_cakismasi_denetle(odalar)

    # Kanonik slot sayısı: sayfalar arasında ortak olan uzunluk.
    uzunluklar = {len(p["slot_labels"]) for p in programlar.values()}
    slot_sayisi = min(uzunluklar) if uzunluklar else 0

    # Görüntülenecek slot etiketi: en sık kullanılan yazım.
    from collections import Counter
    etiket_sayaci = [Counter() for _ in range(slot_sayisi)]
    for p in programlar.values():
        for i, e in enumerate(p["slot_labels"][:slot_sayisi]):
            etiket_sayaci[i][e] += 1
    zaman_slotlari = [
        {"index": i,
         "label": etiket_sayaci[i].most_common(1)[0][0] if etiket_sayaci[i] else f"{i + 1}. saat",
         "variants": sorted(etiket_sayaci[i])}
        for i in range(slot_sayisi)
    ]

    kayitlar: List[dict] = []
    eslesen, eslesmeyen = [], []
    for o in sorted(odalar, key=lambda x: (x["floor"], x["room_key"])):
        p = programlar.get(o["room_key"])
        kayit = dict(o)
        if p:
            eslesen.append({"room_code": o["room_code"], "sheet": p["sheet_name"]})
            kayit["has_schedule"] = True
            kayit["sheet_name"] = p["sheet_name"]
            kayit["slot_labels"] = p["slot_labels"][:slot_sayisi]
            kayit["schedule"] = {g: (p["grid"][g] + [None] * slot_sayisi)[:slot_sayisi]
                                 for g in GUNLER}
        else:
            # PROGRAMI YOK — sıfır dolulukla DOLDURULMAZ. "Veri yok"tur.
            eslesmeyen.append({"room_code": o["room_code"], "floor": o["floor"],
                               "reason": "xlsx_icinde_program_sayfasi_yok"})
            kayit["has_schedule"] = False
            kayit["schedule"] = None
        kayitlar.append(kayit)

    envanter_anahtarlari = {o["room_key"] for o in odalar}
    sahipsiz = [{"sheet": p["sheet_name"], "room_key": k,
                 "reason": "KAT_PLANI_envanterinde_yok"}
                for k, p in sorted(programlar.items())
                if k not in envanter_anahtarlari]

    veri = {
        "meta": {
            "source_workbook": XLSX.name,
            "source_plans": [k["pdf"] for k in KAT_KAYNAKLARI],
            "semester": "2025-2026 Bahar",
            "days": GUNLER,
            "slot_count": slot_sayisi,
            "generated_by": "build_infrastructure_map.py",
            "read_only": True,
            "note": ("Ham dosyalardan türetilmiştir; veritabanına yazılmaz. "
                     "Boş slot kaynakta '0' ile gösterilir ve burada null olur."),
        },
        "floors": [
            {**k, "room_count": sum(1 for o in odalar if o["floor"] == k["floor"])}
            for k in KAT_KAYNAKLARI
        ],
        "time_slots": zaman_slotlari,
        "rooms": kayitlar,
    }

    rapor = {
        "total_rooms_in_inventory": len(odalar),
        "total_schedule_sheets": len(programlar),
        "matched_count": len(eslesen),
        "unmatched_rooms_without_schedule": eslesmeyen,
        "schedule_sheets_without_inventory": sahipsiz,
        "normalization_collisions": catismalar,
        "rejected_non_room_values": elenenler,
        "slot_label_variants": [s for s in zaman_slotlari if len(s["variants"]) > 1],
        "floor_mapping_evidence": {
            "method": "PDF kot (yükseklik) sırası + AMFİ etiketlerinin kat teyidi",
            "detail": [{"pdf": k["pdf"], "internal_title": k["internal_title"],
                        "elevation_m": k["elevation_m"], "mapped_floor": k["floor"]}
                       for k in KAT_KAYNAKLARI],
        },
    }
    return veri, rapor


def plan_konumlarini_cikar(oda_anahtarlari: Dict[str, int]) -> dict:
    """PDF planlarındaki oda ETİKETLERİNDEN doğrulanabilir konum çıkarır.

    NEDEN ÇOĞU ODA İÇİN KONUM YOK
    -----------------------------
    Planlar mimari çizimdir: odaların içine Excel'deki kod ("C 208")
    değil, mimarın sıra numarası ve alanı ("13-86.7") yazılmıştır. Bu
    sıra numarasını Excel koduna eşlemek TAHMİN olurdu ve yanlış odayı
    boyayabilirdi. Bu yüzden yalnızca kaynakta AÇIKÇA oda kodu yazan
    etiketler konumlandırılır; gerisi `unplaced` kalır ve arayüzde plan
    dışında, ama tam işlevsel bir ızgarada gösterilir.

    Koordinatlar 0–1 aralığında NORMALİZE edilir: arka plan görseli
    yeniden ölçeklendiğinde işaretler yerini kaybetmez.
    """
    try:
        import pdfplumber
    except ImportError:
        return {"rooms": {}, "note": "pdfplumber kurulu değil; konum çıkarılmadı."}

    konumlar: Dict[str, dict] = {}
    for kaynak in KAT_KAYNAKLARI:
        yol = HAM / kaynak["pdf"]
        if not yol.exists():
            continue
        with pdfplumber.open(yol) as pdf:
            sayfa = pdf.pages[0]
            gen, yuk = float(sayfa.width), float(sayfa.height)
            for w in sayfa.extract_words():
                anahtar = oda_anahtari(w["text"])
                if oda_anahtarlari.get(anahtar) != kaynak["floor"]:
                    continue
                konumlar[anahtar] = {
                    "floor": kaynak["floor"],
                    "x": round((float(w["x0"]) + float(w["x1"])) / 2 / gen, 4),
                    "y": round((float(w["top"]) + float(w["bottom"])) / 2 / yuk, 4),
                    "source": f"{kaynak['pdf']} · plan etiketi \"{w['text']}\"",
                    "confidence": "label_exact",
                }
    return {
        "coordinate_system": ("Arka plan görselinin sol-üst köşesine göre "
                              "0–1 normalize; responsive ölçekte kayma olmaz."),
        "how_to_extend": ("Bir odayı plana yerleştirmek için buraya "
                          "\"<ROOM_KEY>\": {\"floor\":n,\"x\":0-1,\"y\":0-1,"
                          "\"source\":\"kim/nasıl belirledi\","
                          "\"confidence\":\"manual\"} ekleyin. "
                          "TAHMİNLE koordinat girmeyin; yanlış oda boyanır."),
        "rooms": dict(sorted(konumlar.items())),
    }


def main() -> None:
    # Ham kaynaklar TEK yerde (newversion/…/raw) tutulur; kök kopyaya
    # çoğaltılmaz ki iki ayrı "authoritative" nüsha oluşmasın. Bu script
    # kök kopyadan çalıştırılırsa kaynak bulunmaz ve sessizce boş çıktı
    # üretmek yerine açıkça söyler.
    if not XLSX.exists():
        raise SystemExit(
            f"Ham kaynak bulunamadı: {XLSX}\n"
            "Bu üretici, ham dosyaların bulunduğu kopyadan çalıştırılmalıdır "
            "(newversion/integration/backend). Üretilen JSON'lar daha sonra "
            "kök kopyaya kopyalanır."
        )
    TURETILMIS.mkdir(parents=True, exist_ok=True)
    veri, rapor = uret()

    # Plan konumları: yalnızca kanıtlanabilir olanlar.
    kat_haritasi = {r["room_key"]: r["floor"] for r in veri["rooms"]}
    geometri = plan_konumlarini_cikar(kat_haritasi)
    (TURETILMIS / "room_geometry.json").write_text(
        json.dumps(geometri, ensure_ascii=False, indent=1), encoding="utf-8")
    yerlesen = set(geometri["rooms"])
    rapor["plan_placement"] = {
        "placed_count": len(yerlesen),
        "unplaced_count": len(veri["rooms"]) - len(yerlesen),
        "placed": sorted(yerlesen),
        "reason_unplaced": ("Plan PDF'lerinde oda içi etiketler mimari sıra "
                            "numarası + alan ('13-86.7') biçimindedir; Excel "
                            "oda koduna eşlemek tahmin olurdu."),
    }
    print(f"plana yerleşen oda  : {len(yerlesen)}")
    (TURETILMIS / "classroom_usage.json").write_text(
        json.dumps(veri, ensure_ascii=False, indent=1, sort_keys=False), encoding="utf-8")
    (TURETILMIS / "room_match_report.json").write_text(
        json.dumps(rapor, ensure_ascii=False, indent=1, sort_keys=False), encoding="utf-8")
    print(f"envanter odası      : {rapor['total_rooms_in_inventory']}")
    print(f"program sayfası     : {rapor['total_schedule_sheets']}")
    print(f"eşleşen             : {rapor['matched_count']}")
    print(f"programsız oda      : {len(rapor['unmatched_rooms_without_schedule'])}")
    print(f"envantersiz sayfa   : {len(rapor['schedule_sheets_without_inventory'])}")
    print(f"anahtar çakışması   : {len(rapor['normalization_collisions'])}")
    print(f"slot sayısı         : {veri['meta']['slot_count']}")


if __name__ == "__main__":
    main()
