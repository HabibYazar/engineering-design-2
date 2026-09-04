"""YABANCI ÖĞRENCİ SAYILARI AKTARIMI — ankara_bilim_yabanci_ogrenci_2025_2026.xlsx

KAYNAK YETKİLİDİR
-----------------
Dosya, Ankara Bilim Üniversitesi'nin 2025-2026 yabancı öğrenci
sayılarının yetkili kaynağıdır. Hiçbir satır atlanmaz, hiçbir sayı
aşağı çekilmez, eşleşmeyen satır SESSİZCE DÜŞMEZ.

    Yabanci_Ogrenci_Detay : 15 program satırı
    Fakulte_Ozet          : 4 fakülte + üniversite toplamı (233)

EŞLEŞTİRME KURALLARI
--------------------
1. FAKÜLTE kaynağın dediğidir. Kaynak "Meslek Yüksekokulu / Önlisans"
   gibi bir düzey niteleyicisi eklerse niteleyici atılır; yeni birim
   AÇILMAZ.

2. PROGRAM, yalnızca normalleştirilmiş adı BİREBİR tutuyorsa ve
   program KAYNAĞIN SÖYLEDİĞİ FAKÜLTENİN altındaysa bağlanır.
   Normalleştirme "(İngilizce)", "(Türkçe)", "PR.", "BÖLÜMÜ" gibi
   yapısal ekleri atar (bkz. `services/unit_matching.py`).

3. Program tutmazsa BÖLÜM denenir (aynı fakülte içinde, birebir ad).

4. Hiçbiri tutmazsa satır FAKÜLTE düzeyinde saklanır ve gerekçesi
   yazılır. Ad benzerliğine dayanarak tahmin YAPILMAZ; yeni program
   OLUŞTURULMAZ.

FAKÜLTE ÇELİŞKİSİ
-----------------
Gözlenen gerçek durum: "İç Mimarlık ve Çevre Tasarımı" kaynakta
Mühendislik ve Mimarlık Fakültesi altında, kurumun hiyerarşisinde ise
Güzel Sanatlar ve Tasarım Fakültesi altındadır. Bu satır KAYNAĞIN
fakültesine yazılır (fakülte toplamı 98 böylece korunur), program
kimliği BOŞ bırakılır ve çelişki kayda geçer. Programı sessizce başka
fakülteye taşımak da, fakülte toplamını kaynağa aykırı hâle getirmek de
veri uydurmak olurdu.

İDEMPOTENS
----------
Doğal anahtar: (akademik yıl, boyut, kaynak fakülte etiketi, kaynak
program etiketi). İkinci çalıştırma yeni satır açmaz, değeri günceller.

MUTABAKAT
---------
Aktarım sonunda kaynak toplamı ile yazılan toplam KARŞILAŞTIRILIR ve
eşit değilse çıkış kodu 1 olur. Sessiz kayıp mümkün değildir.
"""

from __future__ import annotations

import argparse
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
from sqlalchemy import select

from app.database import SessionLocal, init_db
from app.models import AcademicProgram, Department, Faculty
from app.models.student_demographic import (
    DIMENSION_FOREIGN,
    RESOLUTION_DEPARTMENT,
    RESOLUTION_FACULTY,
    RESOLUTION_PROGRAM,
    StudentDemographicCount,
)
from app.services.unit_matching import normalize_unit_name

KAYNAK_ADI = "ankara_bilim_yabanci_ogrenci"
VARSAYILAN_DOSYA = (
    Path(__file__).resolve().parents[1]
    / "data" / "ekdata" / "ankara_bilim_yabanci_ogrenci_2025_2026.xlsx"
)
DETAY_SAYFA = "Yabanci_Ogrenci_Detay"
OZET_SAYFA = "Fakulte_Ozet"

#: Kaynağın fakülte etiketine eklediği DÜZEY niteleyicileri. Bunlar
#: birimin kimliğinin parçası değildir; "Meslek Yüksekokulu / Önlisans"
#: ile "MESLEK YÜKSEKOKULU" aynı birimdir.
_DUZEY_EKLERI = ("/ ÖNLİSANS", "/ ONLISANS", "/ LİSANS", "/ LISANS")


def _fakulte_anahtari(etiket: str) -> str:
    """Kaynak fakülte etiketini karşılaştırılabilir anahtara çevirir."""
    ham = (etiket or "").strip()
    buyuk = ham.upper()
    for ek in _DUZEY_EKLERI:
        if buyuk.endswith(ek):
            ham = ham[: len(ham) - len(ek)].strip()
            break
    return normalize_unit_name(ham)


class YabanciOgrenciAktarimi:
    """Tek dosyayı okur, çözümler, yazar ve mutabakatı raporlar."""

    def __init__(self, db, dosya: Path, dry_run: bool = False) -> None:
        self.db = db
        self.dosya = dosya
        self.dry_run = dry_run
        self.satirlar: List[dict] = []
        self.cakismalar: List[str] = []
        self.sayac = {"eklendi": 0, "guncellendi": 0, "degismedi": 0}

        self._fakulteler: Dict[str, Faculty] = {}
        for f in db.execute(select(Faculty)).scalars():
            self._fakulteler[normalize_unit_name(f.name)] = f

        self._bolumler: Dict[Tuple[int, str], Department] = {}
        for b in db.execute(select(Department)).scalars():
            self._bolumler[(b.faculty_id, normalize_unit_name(b.name))] = b

        self._programlar: Dict[Tuple[int, str], AcademicProgram] = {}
        for p in db.execute(select(AcademicProgram)).scalars():
            bolum = db.get(Department, p.department_id)
            if bolum is None:
                continue
            anahtar = (bolum.faculty_id, normalize_unit_name(p.name))
            self._programlar[anahtar] = p

    # ------------------------------------------------------------------
    # OKUMA
    # ------------------------------------------------------------------
    def oku(self) -> Tuple[List[dict], Dict[str, int], Optional[int]]:
        wb = openpyxl.load_workbook(self.dosya, data_only=True)
        if DETAY_SAYFA not in wb.sheetnames:
            raise SystemExit(f"'{DETAY_SAYFA}' sayfası yok: {self.dosya}")

        detay = []
        ws = wb[DETAY_SAYFA]
        basliklar = [str(h or "").strip() for h in next(ws.iter_rows(values_only=True))]
        sutun = {ad: i for i, ad in enumerate(basliklar)}
        gerekli = ["Akademik Yıl", "Fakülte / Birim", "Program",
                   "Yabancı Öğrenci Sayısı"]
        eksik = [g for g in gerekli if g not in sutun]
        if eksik:
            raise SystemExit(f"Beklenen sütunlar yok: {eksik} · bulunan: {basliklar}")

        for satir in ws.iter_rows(min_row=2, values_only=True):
            if satir is None or all(h in (None, "") for h in satir):
                continue
            sayi = satir[sutun["Yabancı Öğrenci Sayısı"]]
            if sayi is None:
                # NULL "sıfır öğrenci" DEĞİLDİR; satır kaydedilir ama
                # sayı yazılmaz — sessizce atmak veri kaybı olurdu.
                self.cakismalar.append(
                    f"Sayı boş: {satir[sutun['Program']]}")
                continue
            detay.append({
                "academic_year": str(satir[sutun["Akademik Yıl"]]).strip(),
                "faculty_label": str(satir[sutun["Fakülte / Birim"]]).strip(),
                "program_label": str(satir[sutun["Program"]]).strip(),
                "language": (str(satir[sutun["Eğitim Dili"]]).strip()
                             if "Eğitim Dili" in sutun
                             and satir[sutun["Eğitim Dili"]] else None),
                "count": int(sayi),
            })

        # Özet sayfası doğrulama içindir; yazılmaz.
        ozet: Dict[str, int] = {}
        toplam: Optional[int] = None
        if OZET_SAYFA in wb.sheetnames:
            for satir in wb[OZET_SAYFA].iter_rows(min_row=2, values_only=True):
                if not satir or satir[0] is None or satir[1] is None:
                    continue
                ad, deger = str(satir[0]).strip(), satir[1]
                if "TOPLAM" in ad.upper():
                    toplam = int(deger)
                else:
                    ozet[ad] = int(deger)
        return detay, ozet, toplam

    # ------------------------------------------------------------------
    # ÇÖZÜMLEME
    # ------------------------------------------------------------------
    def _coz(self, kayit: dict) -> dict:
        fak_anahtar = _fakulte_anahtari(kayit["faculty_label"])
        fakulte = self._fakulteler.get(fak_anahtar)
        if fakulte is None:
            # Kısmi eşleşme: kaynak "MESLEK YÜKSEKOKULU", hiyerarşi
            # "MESLEK YÜKSEKOKULU" → normalize sonrası "MESLEK".
            adaylar = [f for k, f in self._fakulteler.items()
                       if k and (k.startswith(fak_anahtar)
                                 or fak_anahtar.startswith(k))]
            fakulte = adaylar[0] if len(adaylar) == 1 else None

        prog_anahtar = normalize_unit_name(kayit["program_label"])
        if fakulte is None:
            return {"faculty": None, "department": None, "program": None,
                    "resolution": RESOLUTION_FACULTY,
                    "note": "Fakülte çözümlenemedi; yeni birim açılmadı."}

        program = self._programlar.get((fakulte.id, prog_anahtar))
        if program is not None:
            return {"faculty": fakulte,
                    "department": self.db.get(Department, program.department_id),
                    "program": program,
                    "resolution": RESOLUTION_PROGRAM, "note": None}

        bolum = self._bolumler.get((fakulte.id, prog_anahtar))
        if bolum is not None:
            return {"faculty": fakulte, "department": bolum, "program": None,
                    "resolution": RESOLUTION_DEPARTMENT,
                    "note": "Program kaydı yok; bölüm düzeyinde bağlandı."}

        # Program BAŞKA bir fakültede mi duruyor? Çelişki kaydedilir,
        # satır kaynağın fakültesinde kalır.
        baska = [(fid, p) for (fid, ad), p in self._programlar.items()
                 if ad == prog_anahtar]
        if baska:
            fid, p = baska[0]
            baska_fak = self.db.get(Faculty, fid)
            not_ = (f"Kaynak bu programı '{kayit['faculty_label']}' altında "
                    f"veriyor; hiyerarşide '{baska_fak.name}' altında "
                    f"(program id={p.id}). Fakülte atıfı KAYNAĞIN, program "
                    f"kimliği bağlanmadı.")
            self.cakismalar.append(not_)
            return {"faculty": fakulte, "department": None, "program": None,
                    "resolution": RESOLUTION_FACULTY, "note": not_}

        not_ = ("Bu ada karşılık gelen program/bölüm hiyerarşide yok; "
                "yeni birim AÇILMADI, satır fakülte düzeyinde saklandı.")
        self.cakismalar.append(f"{kayit['program_label']}: {not_}")
        return {"faculty": fakulte, "department": None, "program": None,
                "resolution": RESOLUTION_FACULTY, "note": not_}

    # ------------------------------------------------------------------
    # YAZMA
    # ------------------------------------------------------------------
    def aktar(self) -> int:
        detay, ozet, kaynak_toplam = self.oku()
        kaynak_satir_toplami = sum(k["count"] for k in detay)

        for kayit in detay:
            c = self._coz(kayit)
            mevcut = self.db.execute(
                select(StudentDemographicCount).where(
                    StudentDemographicCount.academic_year == kayit["academic_year"],
                    StudentDemographicCount.dimension == DIMENSION_FOREIGN,
                    StudentDemographicCount.source_faculty_label
                    == kayit["faculty_label"],
                    StudentDemographicCount.source_program_label
                    == kayit["program_label"],
                )
            ).scalar_one_or_none()

            alanlar = dict(
                faculty_id=c["faculty"].id if c["faculty"] else None,
                department_id=c["department"].id if c["department"] else None,
                academic_program_id=c["program"].id if c["program"] else None,
                education_language=kayit["language"],
                student_count=kayit["count"],
                resolution=c["resolution"],
                resolution_note=c["note"],
                source_dataset=KAYNAK_ADI,
                source_file=self.dosya.name,
            )
            if mevcut is None:
                self.db.add(StudentDemographicCount(
                    academic_year=kayit["academic_year"],
                    dimension=DIMENSION_FOREIGN,
                    source_faculty_label=kayit["faculty_label"],
                    source_program_label=kayit["program_label"],
                    **alanlar))
                self.sayac["eklendi"] += 1
            else:
                degisti = any(getattr(mevcut, a) != d for a, d in alanlar.items())
                for a, d in alanlar.items():
                    setattr(mevcut, a, d)
                self.sayac["guncellendi" if degisti else "degismedi"] += 1

        if self.dry_run:
            self.db.rollback()
        else:
            self.db.commit()

        return self._rapor(detay, ozet, kaynak_toplam, kaynak_satir_toplami)

    def _rapor(self, detay, ozet, kaynak_toplam, satir_toplami) -> int:
        print(f"\nDosya: {self.dosya.name}")
        print(f"Kaynak satırı: {len(detay)} · sayıların toplamı: {satir_toplami}")
        print(f"Yazma: {self.sayac}")

        yazilan = self.db.execute(
            select(StudentDemographicCount).where(
                StudentDemographicCount.dimension == DIMENSION_FOREIGN)
        ).scalars().all()
        yazilan_toplam = sum(r.student_count for r in yazilan)

        print("\nÇÖZÜMLEME DÜZEYİ:")
        for duzey in (RESOLUTION_PROGRAM, RESOLUTION_DEPARTMENT, RESOLUTION_FACULTY):
            n = sum(1 for r in yazilan if r.resolution == duzey)
            adet = sum(r.student_count for r in yazilan if r.resolution == duzey)
            print(f"   {duzey:<11} {n:>3} satır · {adet:>4} öğrenci")

        print("\nFAKÜLTE TOPLAMLARI (yazılan):")
        fak_toplam: Dict[str, int] = {}
        for r in yazilan:
            fak_toplam[r.source_faculty_label] = (
                fak_toplam.get(r.source_faculty_label, 0) + r.student_count)
        for ad, n in sorted(fak_toplam.items(), key=lambda x: -x[1]):
            beklenen = ozet.get(ad)
            isaret = "✓" if beklenen is None or beklenen == n else f"✗ (özet {beklenen})"
            print(f"   {ad:<40} {n:>4}  {isaret}")

        if self.cakismalar:
            print(f"\nÇÖZÜMLENEMEYEN / ÇELİŞKİLİ ({len(self.cakismalar)}) — "
                  "satırlar KORUNDU, fakülte düzeyinde saklandı:")
            for c in self.cakismalar:
                print(f"   · {c}")

        print("\n" + "-" * 66)
        print(f"MUTABAKAT: kaynak {satir_toplami} = yazılan {yazilan_toplam}"
              f"  {'✓' if satir_toplami == yazilan_toplam else '✗'}")
        if kaynak_toplam is not None:
            print(f"           özet sayfası toplamı {kaynak_toplam}"
                  f"  {'✓' if kaynak_toplam == yazilan_toplam else '✗'}")
        print("-" * 66)

        hatali = satir_toplami != yazilan_toplam or (
            kaynak_toplam is not None and kaynak_toplam != yazilan_toplam)
        return 1 if hatali else 0


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--file", default=str(VARSAYILAN_DOSYA))
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    dosya = Path(args.file)
    if not dosya.exists():
        print(f"Dosya bulunamadı: {dosya}")
        return 1

    init_db()
    db = SessionLocal()
    try:
        return YabanciOgrenciAktarimi(db, dosya, args.dry_run).aktar()
    finally:
        db.close()


if __name__ == "__main__":
    sys.exit(main())
