#!/usr/bin/env python
"""PART3 AKTARIMI — derslik envanteri + eğitim ücretleri.

KAYNAK (data/ekdata/part3)
--------------------------
  1. ABU_Tum_Derslikler.xlsx
     Kampüsün derslik/laboratuvar envanteri. Sayfa ÜÇ KAT BLOĞU hâlinde
     yatay düzenlenmiş: A-E = kat 0, G-K = kat 1, M-Q = kat 2. Her blok
     (kod, etiket, sınıf kapasitesi, öğrenci kapasitesi, sahip birim)
     taşır. Blok sonlarında ara toplam, sağ altta birim başına oda
     sayısı özeti, altta ders planlama kuralları metni bulunur.

  2. Ankara_Bilim_Universitesi_Egitim_Ucretleri (1).xlsx
     Kendi programlarımızın yıllık ücreti. Üç sayfa:
       · "2026-2027 Eğitim Ücretleri"  (dil program adına gömülü,
                                        ayrıca ilk-5-tercih ve peşin
                                        ödeme sütunları)
       · "2025-2026 Ücretleri"         (dil AYRI sütun — en zengin biçim)
       · "2025-2026 Eğitim Ücretleri"  (aynı yıl, dil program adına gömülü)

  3. ankara_ozel_universiteler_10_egitim_ucretleri.xlsx
     10 rakip vakıf üniversitesinin ücretleri; tek şemada birleştirilmiş.
     "Genel Ozet" sayfası MINIFS/MAXIFS FORMÜLLERİDİR (dosyada değer
     yoktur) ve TÜRETİLMİŞTİR — aktarılmaz, gerekirse sorgudan üretilir.
     "Notlar" sayfası açıklama metnidir.

PART3 YETKİLİDİR
----------------
Çakışma hâlinde part3 kazanır. Şu an fiilen çakışan bir alan yoktur:
derslik envanteri boş bir tabloyu doldurur, ücretler ise şemada hiç
bulunmayan bir kavramdır. Yine de aktarım, mevcut bir değeri
değiştirdiğinde bunu `data_source_conflicts` üzerine
`resolution="applied_incoming"` ile yazar — hangi sayının neyi ezdiği
kayıt altında kalsın.

ELEME YOK
---------
Eşleşmeyen satır ATILMAZ. Program kimliğine çözülemeyen ücret satırı
bölüm/fakülte kimliğiyle, o da yoksa ham adıyla saklanır ve raporda
gerekçesiyle listelenir. Sahibi akademik bir fakülteye çözülemeyen
derslik (Hazırlık Okulu, Müzik) `owner_label` ile korunur.

İDEMPOTENT
----------
Her yazma doğal anahtar üzerinden upsert'tir:
    derslik  : physical_facilities.code
    ücret    : (yıl, program, dil, ücret türü)
    rakip    : (kurum, yıl, seviye, birim, program, etiket, ücret metni)

KULLANIM
--------
    python import_part3.py
    python import_part3.py --dry-run
    python import_part3.py --dir <yol>
"""

from __future__ import annotations

import argparse
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

BACKEND_DIR = Path(__file__).resolve().parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

from sqlalchemy import select  # noqa: E402

from app.database import SessionLocal, init_db  # noqa: E402
from app.models import (  # noqa: E402
    AcademicProgram,
    BenchmarkInstitution,
    CompetitorTuitionFee,
    DataSourceConflict,
    Department,
    Faculty,
    PhysicalFacility,
    ProgramTuitionFee,
)
from app.models.tuition_fee import (  # noqa: E402
    FEE_FULL,
    FEE_FULL_SCHOLARSHIP,
    FEE_HALF_SCHOLARSHIP,
    FEE_OTHER_DISCOUNT,
    LEVEL_ASSOCIATE,
    LEVEL_BACHELOR,
    LEVEL_HEALTH,
    LEVEL_PREP,
)

# --------------------------------------------------------------------------
# Yol çözümü — önce integration içi, sonra depo kökü
# --------------------------------------------------------------------------
#: Taze bir `git clone` yalnızca `integration/` klasörünü kurabilsin diye
#: veri artık depo içine kopyalanmıştır. Eski yerleşim de destekleniyor
#: ki mevcut çalışma kopyaları bozulmasın.
ADAY_DIZINLER = (
    BACKEND_DIR.parent / "data" / "ekdata" / "part3",       # integration/data
    BACKEND_DIR.parent.parent / "data" / "ekdata" / "part3",  # depo kökü
)


def varsayilan_dizin() -> Path:
    for aday in ADAY_DIZINLER:
        if aday.exists():
            return aday
    return ADAY_DIZINLER[0]


KAYNAK_KUMESI = "part3"

_TR = str.maketrans("ÇĞİIÖŞÜçğıiöşü", "CGIIOSUcgiiosu")


def kat(metin: Optional[str]) -> str:
    """Türkçe duyarlı katlama — YALNIZCA eşleştirme için."""
    if not metin:
        return ""
    d = unicodedata.normalize("NFKD", str(metin).translate(_TR))
    return " ".join(
        "".join(c for c in d if not unicodedata.combining(c)).upper().split())


def _hucre(ws, r: int, c: int) -> str:
    v = ws.cell(r, c).value
    return "" if v is None else str(v).strip()


def _tam_sayi(metin: str) -> Optional[int]:
    """Hücreyi tam sayıya çevirir. Boş/sayısal olmayan → None (0 DEĞİL)."""
    metin = (metin or "").strip()
    if not metin:
        return None
    try:
        return int(float(metin.replace(".", "").replace(",", ".")))
    except ValueError:
        return None


def _para(metin) -> Optional[Decimal]:
    """"928000" / "525.000 TL" → Decimal. Aralık metni → None."""
    if metin is None:
        return None
    m = str(metin).strip()
    if not m or m == "-":
        return None
    # Aralık ("386.000 TL - 410.000 TL") sayıya indirgenmez.
    if re.search(r"\d\s*(TL)?\s*[-–]\s*\d", m):
        return None
    temiz = re.sub(r"[^\d,.]", "", m)
    if not temiz:
        return None
    # Türkçe biçim: nokta binlik, virgül ondalık.
    temiz = temiz.replace(".", "").replace(",", ".")
    try:
        return Decimal(temiz)
    except InvalidOperation:
        return None


# ==========================================================================
# 1) DERSLİK ENVANTERİ
# ==========================================================================

#: Kaynak sayfadaki üç kat bloğunun başlangıç sütunu ve kat numarası.
KAT_BLOKLARI: Tuple[Tuple[int, int], ...] = ((1, 0), (7, 1), (13, 2))

#: Sahiplik kısaltması → fakülte KODU.
#:
#: Bu bir tahmin değil, kurumun kendi kısaltma listesidir; sağ alttaki
#: özet bloğu da aynı kısaltmaları kullanır. "ITBF" (Türkçe harfsiz)
#: ile "İTBF" AYNI fakültedir; ayrı bırakmak fakülteyi ikiye bölerdi.
SAHIP_FAKULTE: Dict[str, str] = {
    "MMF": "MUHMIM",
    "MMF-LAB": "MUHMIM",
    "ITBF": "INSTOPBIL",
    "GSTF": "GUZSANTAS",
    "HF": "HUKUK",
    "HUK": "HUKUK",
    "MYO": "MESLEK",
}

#: Akademik fakülteye BAĞLANMAYAN ama gerçek olan sahipler. Kayıt
#: dışlanmaz; `owner_label` ile korunur ve üniversite kapsamında görünür.
FAKULTE_DISI_SAHIPLER = {"HAZIRLIK OKULU", "MUZIK"}

#: Sağ alttaki "birim → oda sayısı" özet bloğunun satırları oda DEĞİLDİR.
OZET_SATIR_KODLARI = {kat(k) for k in SAHIP_FAKULTE} | {
    kat(k) for k in FAKULTE_DISI_SAHIPLER} | {"TOPLAM", "TOPLAM:"}


def _derslik_turu(kod: str, etiket: str, sahip: str) -> str:
    """Mekân türü — kaynağın kendi işaretlerinden."""
    k, e, s = kat(kod), kat(etiket), kat(sahip)
    if k.startswith("L ") or k.startswith("L") and k[1:2].isdigit():
        return "laboratory"
    if "LAB" in e or "LAB" in s or "STD" in e or "STUD" in e:
        return "laboratory"
    return "classroom"


#: Gerçek bir mekân kodu. POZİTİF kalıp kullanılır: sayfada oda
#: satırlarının yanı sıra başlık ("Ders Planlama Esasları"), kural metni
#: ve birim özeti de bulunuyor. "Oda olmayanı ele" yerine "oda olanı seç"
#: demek, sayfaya yeni bir açıklama eklendiğinde de doğru kalır.
_ODA_KODU = re.compile(
    r"^(?:[A-ZÇĞİÖŞÜ]{1,2}\s?\d{2,4}"       # C 003 · L008 · C071
    r"|AMF[İI]\s?\d+"                       # AMFİ4 · AMFİ 5
    r"|LECTURE\s+HALL\s*\d+)$",             # LECTURE HALL 5
    re.IGNORECASE)


def _oda_mu(kod: str) -> bool:
    return bool(_ODA_KODU.match((kod or "").strip()))


def _oda_kodu(ham: str) -> str:
    """"C 071" ve "C071" AYNI odadır; kod boşluksuz normalleştirilir."""
    return re.sub(r"\s+", "", ham).upper()


class DerslikAktarimi:
    """Envanteri `physical_facilities` üzerine yazar."""

    def __init__(self, db, dry_run: bool = False) -> None:
        self.db = db
        self.dry_run = dry_run
        self.eklendi = 0
        self.guncellendi = 0
        self.degismedi = 0
        self.satirlar: List[dict] = []
        self.sahip_dagilimi: Counter = Counter()
        self.fakultesiz: List[str] = []
        self.cakismalar: List[dict] = []

    def oku(self, yol: Path) -> List[dict]:
        import openpyxl

        ws = openpyxl.load_workbook(yol, data_only=True).worksheets[0]
        odalar: List[dict] = []
        for c0, kat_no in KAT_BLOKLARI:
            for r in range(4, ws.max_row + 1):
                kod = _hucre(ws, r, c0)
                if not kod:
                    continue
                if not _oda_mu(kod):
                    # Birim özeti, başlık ve ders planlama kuralı metni.
                    continue
                odalar.append({
                    "floor": kat_no,
                    "code": _oda_kodu(kod),
                    "source_code": kod,
                    "room_label": _hucre(ws, r, c0 + 1) or None,
                    "capacity": _tam_sayi(_hucre(ws, r, c0 + 2)),
                    "student_capacity": _tam_sayi(_hucre(ws, r, c0 + 3)),
                    "owner": _hucre(ws, r, c0 + 4) or None,
                })
        self.satirlar = odalar
        return odalar

    def aktar(self, odalar: List[dict], dosya: str) -> None:
        fakulteler = {kat(f.code): f for f in
                      self.db.execute(select(Faculty)).scalars()}

        for o in odalar:
            sahip_k = kat(o["owner"])
            self.sahip_dagilimi[o["owner"] or "(belirtilmemiş)"] += 1

            fakulte_kodu = SAHIP_FAKULTE.get(sahip_k)
            fakulte = fakulteler.get(kat(fakulte_kodu)) if fakulte_kodu else None
            if fakulte is None and sahip_k:
                # Kayıt DIŞLANMAZ; sahiplik etiketi korunur.
                self.fakultesiz.append(f"{o['source_code']} → {o['owner']}")

            ad = o["source_code"]
            if o["room_label"]:
                ad = f"{o['source_code']} ({o['room_label']})"

            mevcut = self.db.execute(
                select(PhysicalFacility).where(
                    PhysicalFacility.code == o["code"])
            ).scalar_one_or_none()

            yeni_degerler = {
                "name": ad,
                "facility_type": _derslik_turu(o["source_code"],
                                               o["room_label"] or "",
                                               o["owner"] or ""),
                "faculty_id": fakulte.id if fakulte else None,
                "capacity": o["capacity"],
                "student_capacity": o["student_capacity"],
                "floor": o["floor"],
                "owner_label": o["owner"],
                "room_label": o["room_label"],
                "source_dataset": KAYNAK_KUMESI,
                "source_file": dosya,
                "is_active": True,
            }

            if mevcut is None:
                if not self.dry_run:
                    self.db.add(PhysicalFacility(code=o["code"],
                                                 **yeni_degerler))
                self.eklendi += 1
                continue

            # PART3 YETKİLİDİR: farklı değer geldiğinde YENİSİ yazılır,
            # eskisi çakışma kaydına düşer.
            degisen = {k: (getattr(mevcut, k), v)
                       for k, v in yeni_degerler.items()
                       if getattr(mevcut, k) != v}
            if not degisen:
                self.degismedi += 1
                continue
            for alan, (eski, yeni) in degisen.items():
                if alan in ("capacity", "student_capacity", "faculty_id") \
                        and eski is not None:
                    self.cakismalar.append({
                        "table_name": "physical_facilities",
                        "record_id": mevcut.id, "field_name": alan,
                        "record_label": o["source_code"],
                        "existing_value": str(eski), "existing_source": "mevcut kayıt",
                        "incoming_value": str(yeni), "incoming_source": dosya,
                        "note": "part3 yetkili kaynaktır; gelen değer yazıldı.",
                    })
                if not self.dry_run:
                    setattr(mevcut, alan, yeni)
            self.guncellendi += 1

        if not self.dry_run:
            self.db.flush()


# ==========================================================================
# 2) KENDİ PROGRAMLARIMIZIN ÜCRETLERİ
# ==========================================================================

#: Ham ücret/indirim etiketi → kapalı liste.
def ucret_turu(etiket: str) -> str:
    e = kat(etiket)
    if not e:
        return FEE_OTHER_DISCOUNT
    if "TAM BURS" in e or e == "BURSLU":
        return FEE_FULL_SCHOLARSHIP
    # "%50 Burslu" ve "%50 İndirimli" AYNI kategoridir.
    if "50" in e and ("BURS" in e or "INDIRIM" in e):
        return FEE_HALF_SCHOLARSHIP
    if "UCRETLI" in e or "TAM UCRET" in e:
        return FEE_FULL
    return FEE_OTHER_DISCOUNT


#: Program adına gömülü eğitim dili. Yalnızca PARANTEZ İÇİ etiket dil
#: sayılır: "İngilizce Mütercim ve Tercümanlık" bir program ADIDIR,
#: dil işareti değildir.
_DIL = re.compile(r"\((İngilizce|Türkçe|İng\.?|Tr\.?)\)\s*$", re.IGNORECASE)


def dil_ayikla(program_adi: str) -> Tuple[str, Optional[str]]:
    m = _DIL.search(program_adi or "")
    if not m:
        return (program_adi or "").strip(), None
    dil = m.group(1).lower()
    normal = "İngilizce" if dil.startswith("i̇ng") or dil.startswith("ing") \
        else "Türkçe"
    return _DIL.sub("", program_adi).strip(), normal


#: Programın adındaki, eşleştirmeye girmeyen ekler.
_EK = re.compile(r"\s*\((Özel Yetenek|Uzaktan|II\. Öğretim)\)\s*", re.I)
_SON_EKLER = (" PR.", " PROGRAMI", " ANABILIM DALI", " ANABİLİM DALI")


def program_anahtari(ad: str) -> str:
    """Karşılaştırma için program adını sadeleştirir.

    Tire ve eğik çizgi AYRAÇ sayılır: kaynak "Elektrik Elektronik
    Mühendisliği" yazarken katalog "Elektrik-Elektronik Mühendisliği"
    yazıyor; bu aynı programdır.
    """
    ad = _EK.sub(" ", ad or "")
    k = kat(re.sub(r"[-–/]", " ", ad))
    for ek in (kat(x) for x in _SON_EKLER):
        if k.endswith(ek):
            k = k[: -len(ek)].strip()
    return k


#: KURUMUN KENDİ KISALTMASI. "İHA" = "İnsansız Hava Aracı"; bu bir tahmin
#: değil, aynı programın iki resmî yazımıdır (MYO kataloğu her ikisini de
#: kullanır). Kısaltmayı çözmeden satır eşleşmez ve gereksiz yere
#: "eşleşmedi" raporuna düşerdi.
PROGRAM_TAKMA_ADLARI: Dict[str, str] = {
    "IHA TEKNOLOJISI VE OPERATORLUGU":
        "INSANSIZ HAVA ARACI TEKNOLOJISI VE OPERATORLUGU",
}

#: Ücret dosyasındaki fakülte adı → fakülte KODU.
FAKULTE_ADI_KODU: Dict[str, str] = {
    "HUKUK FAKULTESI": "HUKUK",
    "MUHENDISLIK VE MIMARLIK FAKULTESI": "MUHMIM",
    "INSAN VE TOPLUM BILIMLERI FAKULTESI": "INSTOPBIL",
    "GUZEL SANATLAR VE TASARIM FAKULTESI": "GUZSANTAS",
    "HAVACILIK VE UZAY BILIMLERI FAKULTESI": "HAVUZABIL",
    "MESLEK YUKSEKOKULU": "MESLEK",
}


def _token_esles(a: str, b: str) -> bool:
    """Bağlaçlar atıldıktan sonra belirteç belirtece eşleşiyor mu?

    "FILM TASARIM VE YONETIMI" ile "FILM TASARIMI VE YONETIMI" aynı
    programdır; tek fark iyelik ekidir. Belirteçlerden biri diğerinin
    ÖN EKİ ise eşleşmiş sayılır. Bu, gevşek bir benzerlik ölçüsü
    DEĞİLDİR: sıra ve belirteç sayısı korunmak zorundadır.
    """
    at = [t for t in a.split() if t not in {"VE", "ILE"}]
    bt = [t for t in b.split() if t not in {"VE", "ILE"}]
    if len(at) != len(bt):
        return False
    return all(x == y or x.startswith(y) or y.startswith(x)
               for x, y in zip(at, bt))


class UcretAktarimi:
    """ABÜ program ücretlerini `program_tuition_fees` üzerine yazar."""

    def __init__(self, db, dry_run: bool = False) -> None:
        self.db = db
        self.dry_run = dry_run
        self.eklendi = 0
        self.guncellendi = 0
        self.degismedi = 0
        self.dil_zaten_var = 0
        #: MUTABAKAT. Kaynaktaki her veri satırı şu dört kovadan BİRİNE
        #: düşmek zorundadır; toplamları veri satırı sayısına eşit olmalı.
        #:   eklendi        yeni doğal anahtar
        #:   birlesen       aynı anahtar, AYNI değer (aynı ücretin ikinci sunumu)
        #:   dil_zaten_var  dilsiz sunum, dili bilinen aynı ücret zaten var
        #:   atlanan        ücreti okunamayan satır (gerekçesiyle listelenir)
        self.birlesen = 0
        self.atlanan: List[str] = []
        self.okunan_satir = 0
        self.programsiz: List[str] = []
        self.cakismalar: List[dict] = []
        self._dizin: Optional[Dict[int, List[Tuple[str, AcademicProgram]]]] = None

    # -- eşleştirme --------------------------------------------------

    def _program_dizini(self) -> Dict[int, List[Tuple[str, AcademicProgram]]]:
        """Fakülte kimliği → [(anahtar, program)] — HİYERARŞİ üzerinden."""
        if self._dizin is not None:
            return self._dizin
        bolumler = {d.id: d for d in
                    self.db.execute(select(Department)).scalars()}
        dizin: Dict[int, List[Tuple[str, AcademicProgram]]] = defaultdict(list)
        for p in self.db.execute(select(AcademicProgram)).scalars():
            b = bolumler.get(p.department_id)
            if b is None:
                continue
            dizin[b.faculty_id].append((program_anahtari(p.name), p))
        self._dizin = dizin
        return dizin

    def program_bul(self, fakulte: Optional[Faculty], program_adi: str
                    ) -> Optional[AcademicProgram]:
        """Programı FAKÜLTE İÇİNDE arar; kapsam dışına taşmaz."""
        if fakulte is None:
            return None
        adaylar = self._program_dizini().get(fakulte.id, [])
        aranan = program_anahtari(program_adi)
        aranan = PROGRAM_TAKMA_ADLARI.get(aranan, aranan)

        for anahtar, p in adaylar:
            if anahtar == aranan:
                return p
        # İyelik eki / bağlaç farkı — sıra ve belirteç sayısı korunur.
        esles = [p for anahtar, p in adaylar if _token_esles(anahtar, aranan)]
        # Belirsizlik varsa eşleştirmeyiz: yanlış programa ücret yazmak,
        # eşleşmemiş bırakmaktan daha kötüdür.
        return esles[0] if len(esles) == 1 else None

    # -- aktarım -----------------------------------------------------

    def aktar(self, satirlar: List[dict], dosya: str, sayfa: str) -> None:
        self.okunan_satir += len(satirlar)
        fakulteler = {kat(f.code): f for f in
                      self.db.execute(select(Faculty)).scalars()}
        bolumler = {d.id: d for d in
                    self.db.execute(select(Department)).scalars()}

        for s in satirlar:
            fak_kod = FAKULTE_ADI_KODU.get(kat(s["faculty_name"]))
            fakulte = fakulteler.get(kat(fak_kod)) if fak_kod else None
            program = self.program_bul(fakulte, s["program_name"])
            bolum = bolumler.get(program.department_id) if program else None

            if program is None:
                self.programsiz.append(
                    f"{s['academic_year']} · {s['faculty_name']} · "
                    f"{s['program_name']} "
                    f"({'fakülte çözüldü' if fakulte else 'FAKÜLTE ÇÖZÜLEMEDİ'})")

            anahtar = dict(
                academic_year=s["academic_year"],
                academic_program_id=program.id if program else None,
                education_language=s["language"],
                fee_type=s["fee_type"],
            )
            mevcut = self.db.execute(
                select(ProgramTuitionFee).filter_by(**anahtar)
            ).scalar_one_or_none()

            # Dil belirtilmemiş satır, AYNI ücretin dili bilinen bir
            # sürümü zaten varsa ikinci kez yazılmaz: aynı ücretin iki
            # farklı sunumu tek gerçektir.
            if mevcut is None and s["language"] is None and program is not None:
                dilli = self.db.execute(
                    select(ProgramTuitionFee).where(
                        ProgramTuitionFee.academic_year == s["academic_year"],
                        ProgramTuitionFee.academic_program_id == program.id,
                        ProgramTuitionFee.fee_type == s["fee_type"],
                        ProgramTuitionFee.education_language.isnot(None),
                        ProgramTuitionFee.annual_fee == s["annual_fee"],
                    )
                ).first()
                if dilli is not None:
                    self.dil_zaten_var += 1
                    continue

            degerler = dict(
                department_id=bolum.id if bolum else None,
                faculty_id=fakulte.id if fakulte else None,
                source_faculty_name=s["faculty_name"],
                source_program_name=s["program_name"],
                source_fee_label=s["fee_label"],
                annual_fee=s["annual_fee"],
                first_five_choice_fee=s.get("first_five"),
                upfront_payment_fee=s.get("upfront"),
                additional_fee_note=s.get("extra_note"),
                source_dataset=KAYNAK_KUMESI,
                source_file=dosya,
                source_sheet=sayfa,
            )

            if mevcut is None:
                if not self.dry_run:
                    self.db.add(ProgramTuitionFee(**anahtar, **degerler))
                self.eklendi += 1
                continue

            # AYNI ücretin ikinci sunumu metadata'yı GERİYE ÇEKMEMELİ:
            # daha zengin sayfa (dil sütunlu) önce işlenir, sonraki sayfa
            # yalnızca BOŞ alanları doldurur. Ücretin kendisi part3
            # yetkisiyle her hâlükârda güncellenir.
            degisen = {
                k: (getattr(mevcut, k), v) for k, v in degerler.items()
                if v is not None and getattr(mevcut, k) != v
                and (getattr(mevcut, k) is None or k == "annual_fee")
            }
            if not degisen:
                self.degismedi += 1
                # Aynı anahtar, AYNI değer: aynı ücretin ikinci sunumu.
                self.birlesen += 1
                continue
            if "annual_fee" in degisen:
                eski, yeni = degisen["annual_fee"]
                self.cakismalar.append({
                    "table_name": "program_tuition_fees",
                    "record_id": mevcut.id, "field_name": "annual_fee",
                    "record_label": f"{s['academic_year']} · {s['program_name']}",
                    "existing_value": str(eski), "existing_source": mevcut.source_sheet,
                    "incoming_value": str(yeni), "incoming_source": sayfa,
                    "note": "part3 yetkili kaynaktır; gelen değer yazıldı.",
                })
            if not self.dry_run:
                for alan, (_, yeni) in degisen.items():
                    setattr(mevcut, alan, yeni)
            self.guncellendi += 1

        if not self.dry_run:
            self.db.flush()


def ucret_sayfasi_oku(ws, atlanan: Optional[List[str]] = None) -> List[dict]:
    """"Eğitim Ücretleri" sayfası — dil program adına gömülü.

    HİÇBİR SATIR SESSİZCE DÜŞMEZ. Ücreti okunamayan satır `atlanan`
    listesine gerekçesiyle yazılır; aktarım raporu bunu gösterir ve
    mutabakat kimliği (veri = aktarılan + birleşen + atlanan) tutmazsa
    fark görünür olur.
    """
    satirlar = []
    for r in range(2, ws.max_row + 1):
        yil = _hucre(ws, r, 1)
        if not yil:
            continue
        ham_ad = _hucre(ws, r, 3)
        ad, dil = dil_ayikla(ham_ad)
        ham_ucret = _hucre(ws, r, 5)
        ucret = _para(ham_ucret)
        if ucret is None:
            if atlanan is not None:
                atlanan.append(
                    f"{ws.title} satır {r}: {ham_ad!r} — ücret hücresi "
                    f"sayıya çevrilemedi ({ham_ucret!r})")
            continue
        ek = _hucre(ws, r, 8)
        satirlar.append({
            "academic_year": yil,
            "faculty_name": _hucre(ws, r, 2),
            "program_name": ad,
            "language": dil,
            "fee_label": _hucre(ws, r, 4),
            "fee_type": ucret_turu(_hucre(ws, r, 4)),
            "annual_fee": ucret,
            "first_five": _para(_hucre(ws, r, 6)),
            "upfront": _para(_hucre(ws, r, 7)),
            "extra_note": None if ek in ("", "-") else ek,
        })
    return satirlar


def ucret_sayfasi_oku_dilli(ws, atlanan: Optional[List[str]] = None
                            ) -> List[dict]:
    """"2025-2026 Ücretleri" — eğitim dili AYRI sütun (en zengin biçim).

    Okunamayan ücret sessizce düşmez; bkz. `ucret_sayfasi_oku`.
    """
    satirlar = []
    for r in range(2, ws.max_row + 1):
        yil = _hucre(ws, r, 1)
        if not yil:
            continue
        ham_ucret = _hucre(ws, r, 6)
        ucret = _para(ham_ucret)
        if ucret is None:
            if atlanan is not None:
                atlanan.append(
                    f"{ws.title} satır {r}: {_hucre(ws, r, 3)!r} — ücret "
                    f"hücresi sayıya çevrilemedi ({ham_ucret!r})")
            continue
        satirlar.append({
            "academic_year": yil,
            "faculty_name": _hucre(ws, r, 2),
            "program_name": _hucre(ws, r, 3),
            "language": _hucre(ws, r, 4) or None,
            "fee_label": _hucre(ws, r, 5),
            "fee_type": ucret_turu(_hucre(ws, r, 5)),
            "annual_fee": ucret,
        })
    return satirlar


# ==========================================================================
# 3) RAKİP ÜNİVERSİTE ÜCRETLERİ
# ==========================================================================

SEVIYE_ESLEME = {
    "LISANS": LEVEL_BACHELOR,
    "ON LISANS / MYO": LEVEL_ASSOCIATE,
    "HAZIRLIK": LEVEL_PREP,
    "SAGLIK BILIMLERI (TIP/DIS/ECZACILIK)": LEVEL_HEALTH,
}

KATEGORI_ESLEME = {
    "TAM UCRETLI": FEE_FULL,
    "STANDART INDIRIM (~%50)": FEE_HALF_SCHOLARSHIP,
    "TAM BURSLU (0 TL)": FEE_FULL_SCHOLARSHIP,
}


#: Rakip dosyasındaki kısaltma → `benchmark_institutions` kaydındaki ad.
#: Kurumun kendi resmî kısaltmasıdır (TOBB ETÜ), benzerlik tahmini değil.
KURUM_TAKMA_ADLARI: Dict[str, str] = {
    "TOBB ETU": "TOBB EKONOMI VE TEKNOLOJI UNIVERSITESI",
}


class RakipUcretAktarimi:
    """Rakip kurum ücretlerini `competitor_tuition_fees` üzerine yazar."""

    def __init__(self, db, dry_run: bool = False) -> None:
        self.db = db
        self.dry_run = dry_run
        self.eklendi = 0
        self.degismedi = 0
        self.guncellendi = 0
        self.aralik_metni = 0
        self.kurumsuz: List[str] = []
        self.kurum_dagilimi: Counter = Counter()

    def _kurum_dizini(self) -> Dict[str, BenchmarkInstitution]:
        return {kat(b.name): b for b in
                self.db.execute(select(BenchmarkInstitution)).scalars()}

    def kurum_bul(self, dizin, ad: str) -> Optional[BenchmarkInstitution]:
        """Kaynak kısaltılmış/aksansız ad kullanır; kurum adıyla eşleştirilir."""
        k = kat(ad)
        k = KURUM_TAKMA_ADLARI.get(k, k)
        if k in dizin:
            return dizin[k]
        # "THKU (Turk Hava Kurumu Universitesi)" → parantez içi resmî ad
        m = re.search(r"\(([^)]+)\)", ad or "")
        if m and kat(m.group(1)) in dizin:
            return dizin[kat(m.group(1))]
        # "TOBB ETU" gibi kısaltmalar: kurum adının BELİRTEÇLERİ kaynakta
        # geçiyorsa eşleşir ("TOBB EKONOMI VE TEKNOLOJI UNIVERSITESI").
        cekirdek = k.replace("UNIVERSITESI", "").replace("UNIVERSITY", "").strip()
        if not cekirdek:
            return None
        adaylar = [v for anahtar, v in dizin.items()
                   if cekirdek and cekirdek in anahtar]
        if len(adaylar) == 1:
            return adaylar[0]
        # Baş harf kısaltması ("TOBB ETU" ⊂ "TOBB EKONOMI VE TEKNOLOJI …")
        adaylar = [v for anahtar, v in dizin.items()
                   if all(p in anahtar for p in cekirdek.split())]
        return adaylar[0] if len(adaylar) == 1 else None

    def aktar(self, satirlar: List[dict], dosya: str) -> None:
        dizin = self._kurum_dizini()
        for s in satirlar:
            kurum = self.kurum_bul(dizin, s["university_name"])
            self.kurum_dagilimi[s["university_name"]] += 1
            if kurum is None and s["university_name"] not in self.kurumsuz:
                self.kurumsuz.append(s["university_name"])

            anahtar = dict(
                university_name=s["university_name"],
                academic_year=s["academic_year"],
                level=s["level"],
                unit_name=s["unit_name"],
                program_name=s["program_name"],
                source_fee_label=s["fee_label"],
                fee_text=s["fee_text"],
            )
            mevcut = self.db.execute(
                select(CompetitorTuitionFee).filter_by(**anahtar)
            ).scalar_one_or_none()
            if s["annual_fee"] is None:
                self.aralik_metni += 1

            degerler = dict(
                benchmark_institution_id=kurum.id if kurum else None,
                fee_type=s["fee_type"],
                source_price_category=s["price_category"],
                annual_fee=s["annual_fee"],
                note=s["note"],
                source_dataset=KAYNAK_KUMESI,
                source_file=dosya,
            )
            if mevcut is None:
                if not self.dry_run:
                    self.db.add(CompetitorTuitionFee(**anahtar, **degerler))
                self.eklendi += 1
                continue
            degisen = [k for k, v in degerler.items()
                       if getattr(mevcut, k) != v]
            if not degisen:
                self.degismedi += 1
                continue
            if not self.dry_run:
                for k, v in degerler.items():
                    setattr(mevcut, k, v)
            self.guncellendi += 1
        if not self.dry_run:
            self.db.flush()


def rakip_sayfasi_oku(ws) -> List[dict]:
    satirlar = []
    for r in range(5, ws.max_row + 1):
        uni = _hucre(ws, r, 1)
        if not uni:
            continue
        ham_ucret = _hucre(ws, r, 7)
        if not ham_ucret:
            continue
        kategori = _hucre(ws, r, 9)
        etiket = _hucre(ws, r, 6)
        tur = KATEGORI_ESLEME.get(kat(kategori)) or ucret_turu(etiket)
        not_ = _hucre(ws, r, 8)
        satirlar.append({
            "university_name": uni,
            "academic_year": _hucre(ws, r, 2),
            "level": SEVIYE_ESLEME.get(kat(_hucre(ws, r, 3))),
            "unit_name": _hucre(ws, r, 4) or None,
            "program_name": _hucre(ws, r, 5) or None,
            "fee_label": etiket or None,
            "fee_type": tur,
            "price_category": kategori or None,
            "annual_fee": _para(ham_ucret),
            "fee_text": ham_ucret,
            "note": None if not_ in ("", "-") else not_,
        })
    return satirlar


# ==========================================================================
# Giriş noktası
# ==========================================================================


def _cakismalari_yaz(db, cakismalar: List[dict], dry_run: bool) -> None:
    if dry_run:
        return
    for c in cakismalar:
        var = db.execute(
            select(DataSourceConflict).where(
                DataSourceConflict.table_name == c["table_name"],
                DataSourceConflict.record_id == c["record_id"],
                DataSourceConflict.field_name == c["field_name"],
                DataSourceConflict.incoming_source == c["incoming_source"],
            )
        ).scalar_one_or_none()
        if var is not None:
            var.existing_value = c["existing_value"]
            var.incoming_value = c["incoming_value"]
            continue
        # PART3 YETKİLİ: gelen değer uygulandı.
        db.add(DataSourceConflict(resolution="applied_incoming", **c))


def main(argv: Optional[Iterable[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dir", type=Path, default=varsayilan_dizin())
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args(list(argv) if argv is not None else None)

    if not args.dir.exists():
        print(f"BULUNAMADI: {args.dir}", file=sys.stderr)
        return 2

    print("=" * 68)
    print("PART3 AKTARIMI — derslikler ve eğitim ücretleri")
    print("=" * 68)
    print(f"Kaynak klasör : {args.dir}")
    if args.dry_run:
        print("MOD           : DRY-RUN (yazma yok)")

    import openpyxl

    init_db()
    db = SessionLocal()
    derslik = DerslikAktarimi(db, args.dry_run)
    ucret = UcretAktarimi(db, args.dry_run)
    rakip = RakipUcretAktarimi(db, args.dry_run)
    tanimsiz: List[str] = []

    try:
        dosyalar = sorted(p for p in args.dir.glob("*.xlsx") if p.is_file())
        print(f"\n[1/4] {len(dosyalar)} dosya bulundu")

        for yol in dosyalar:
            wb = openpyxl.load_workbook(yol, data_only=True)
            basliklar = {kat(ws.cell(1, c).value or "")
                         for ws in wb.worksheets for c in range(1, 6)}
            ad = yol.name

            if "SINIF KODU" in {kat(_hucre(ws, 3, 1)) for ws in wb.worksheets}:
                odalar = derslik.oku(yol)
                print(f"      DERSLİK   {ad}  ({len(odalar)} mekân)")
                derslik.aktar(odalar, ad)
            elif "TUM UNIVERSITELER DETAY" in {kat(w.title) for w in wb.worksheets}:
                ws = wb["Tum Universiteler Detay"]
                sat = rakip_sayfasi_oku(ws)
                print(f"      RAKİP     {ad}  ({len(sat)} satır)")
                rakip.aktar(sat, ad)
                # "Genel Ozet" TÜRETİLMİŞ (MINIFS/MAXIFS) — aktarılmaz.
                # "Notlar" açıklama metnidir.
            elif any("AKADEMIK YIL" in b for b in basliklar):
                # En zengin sayfa ÖNCE: dili ayrı sütunda olan.
                sayfalar = sorted(
                    wb.worksheets,
                    key=lambda w: 0 if kat(w.cell(1, 4).value or "") ==
                    "EGITIM DILI" else 1)
                for ws in sayfalar:
                    dilli = kat(ws.cell(1, 4).value or "") == "EGITIM DILI"
                    sat = (ucret_sayfasi_oku_dilli(ws, ucret.atlanan) if dilli
                           else ucret_sayfasi_oku(ws, ucret.atlanan))
                    # Sayfadaki VERİ satırı sayısı (başlık hariç, dolu).
                    ucret.veri_satiri = getattr(ucret, "veri_satiri", 0) + sum(
                        1 for r in range(2, ws.max_row + 1)
                        if _hucre(ws, r, 1))
                    print(f"      ÜCRET     {ad} · {ws.title}  "
                          f"({len(sat)} satır{', dil sütunlu' if dilli else ''})")
                    ucret.aktar(sat, ad, ws.title)
            else:
                tanimsiz.append(ad)
                print(f"      ATLANDI   {ad} — tanınmayan şema")
            wb.close()

        print("\n[2/4] Çakışmalar kaydediliyor…")
        _cakismalari_yaz(db, derslik.cakismalar + ucret.cakismalar, args.dry_run)

        print("[3/4] Kaydediliyor…")
        if args.dry_run:
            db.rollback()
        else:
            db.commit()
        print("[4/4] Tamam.")
    except Exception:
        db.rollback()
        raise
    finally:
        _rapor(derslik, ucret, rakip, tanimsiz)
        db.close()
    return 0


def _rapor(derslik, ucret, rakip, tanimsiz) -> None:
    def blok(baslik: str, satirlar, sinir: int = 30) -> None:
        print("\n" + "-" * 68)
        print(baslik)
        print("-" * 68)
        satirlar = list(satirlar)
        if not satirlar:
            print("  (yok)")
            return
        for s in satirlar[:sinir]:
            print(f"  · {s}")
        if len(satirlar) > sinir:
            print(f"  … ve {len(satirlar) - sinir} satır daha")

    print("\n" + "=" * 68)
    print("ÖZET")
    print("=" * 68)
    print(f"  Derslik      : {derslik.eklendi} eklendi, "
          f"{derslik.guncellendi} güncellendi, {derslik.degismedi} değişmedi")
    kap = sum(o["capacity"] or 0 for o in derslik.satirlar)
    ogr = sum(o["student_capacity"] or 0 for o in derslik.satirlar)
    print(f"                 sınıf kapasitesi {kap}, öğrenci kapasitesi {ogr}")
    veri_satiri = getattr(ucret, "veri_satiri", ucret.okunan_satir)
    hesap = (ucret.eklendi + ucret.guncellendi + ucret.birlesen
             + ucret.dil_zaten_var + len(ucret.atlanan))
    print(f"  ABÜ ücreti   : {ucret.eklendi} eklendi, "
          f"{ucret.guncellendi} güncellendi, {ucret.birlesen} birleşti "
          f"(aynı ücretin ikinci sunumu), "
          f"{ucret.dil_zaten_var} dil tekrarı, "
          f"{len(ucret.atlanan)} okunamadı")
    print(f"                 MUTABAKAT: kaynak veri satırı {veri_satiri} = "
          f"{hesap} (eklendi+güncellendi+birleşti+dil+okunamadı)"
          f"{'  ✓' if veri_satiri == hesap else '  ✗ FARK VAR'}")
    print(f"  Rakip ücreti : {rakip.eklendi} eklendi, "
          f"{rakip.guncellendi} güncellendi, {rakip.degismedi} değişmedi, "
          f"{rakip.aralik_metni} aralık metni (sayısal değil)")

    blok("DERSLİK SAHİPLİK DAĞILIMI",
         [f"{k}: {v}" for k, v in derslik.sahip_dagilimi.most_common()])
    blok("FAKÜLTEYE BAĞLANMAYAN DERSLİKLER (kayıt korundu)",
         derslik.fakultesiz)
    blok("PROGRAMA ÇÖZÜLEMEYEN ÜCRET SATIRLARI", ucret.programsiz)
    blok("ÜCRETİ OKUNAMAYAN SATIRLAR (aktarılmadı)", ucret.atlanan)
    blok("KIYAS KURUMUNA BAĞLANMAYAN ÜNİVERSİTELER", rakip.kurumsuz)
    blok("ÇAKIŞMALAR (part3 kazandı)",
         [f"{c['table_name']}.{c['field_name']} #{c['record_id']} "
          f"({c['record_label']}): {c['existing_value']} → {c['incoming_value']}"
          for c in derslik.cakismalar + ucret.cakismalar])
    blok("TANINMAYAN DOSYALAR", tanimsiz)
    print("\nBetik idempotenttir; tekrar çalıştırmak kayıt eklemez.")


if __name__ == "__main__":
    raise SystemExit(main())
